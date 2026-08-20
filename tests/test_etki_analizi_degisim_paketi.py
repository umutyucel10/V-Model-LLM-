# -*- coding: utf-8 -*-

from copy import deepcopy
import json
from pathlib import Path
import tempfile
import unittest

from openpyxl import load_workbook
from pypdf import PdfReader

import etki_analizi_degisim_paketi as package_logic
import etki_analizi_degisim_raporlama as reporting
import etki_analizi_izlenebilirlik as traceability
import etki_analizi_simulasyon as simulation


SECTIONS = (
    ("TID", "Kullanıcı Gereksinimi", "left"),
    ("SGD", "Sistem Gereksinimi", "left"),
    ("STT", "Alt Sistem Gereksinimi", "left"),
    ("AST", "Alt Sistem Testi", "right"),
    ("SITET", "Sistem Testi", "right"),
    ("KMTD", "Kabul Testi", "right"),
)


def _flat_data():
    return {
        "TID-001": {
            "type": "TID", "ID": "TID-001",
            "content": "Platform toplam maksimum ağırlığı 10 kg değerini aşmamalıdır.",
            "bound_to": "Yok",
        },
        "SGD-001": {
            "type": "SGD", "ID": "SGD-001",
            "content": "Sistemin maksimum ağırlığı 10 kg değerini aşmamalıdır.",
            "bound_to": "TID-001",
        },
        "STT-001": {
            "type": "STT", "ID": "STT-001",
            "content": "Motor alt sistemi 10 kg sistem sınırına ayrılan tahsise uymalıdır.",
            "bound_to": "SGD-001",
        },
        "AST-001": {
            "type": "AST", "ID": "AST-001",
            "content": "Motor alt sistemi ağırlık tahsisi entegrasyon sırasında doğrulanmalıdır.",
            "bound_to": "STT-001",
        },
        "SITET-001": {
            "type": "SITET", "ID": "SITET-001",
            "content": "Sistem toplam ağırlığının 10 kg sınırını aşmadığı ölçülmelidir.",
            "bound_to": "SGD-001",
        },
        "KMTD-001": {
            "type": "KMTD", "ID": "KMTD-001",
            "content": "Platform 10 kg kabul sınırı müşteri kabulünde doğrulanmalıdır.",
            "bound_to": "TID-001",
        },
    }


def _report(flat):
    return traceability.build_traceability_map(
        "Güvenli Güncelleme Test Projesi",
        flat_data=flat,
        document_sections=SECTIONS,
        persist=False,
        check_lm_studio=False,
    )


def _simulation(report):
    request = simulation.ChangeRequest(
        requirement_id="SGD-001",
        current_value="10 kg",
        proposed_value="8 kg",
        reason="Taşıma ve montaj sınırı daraltıldı.",
        requested_by="Değişiklik Kontrol Kurulu",
        change_type=simulation.CHANGE_NUMERIC_LIMIT,
        assumptions=("Kaynak belge sürümleri günceldir.",),
        query="SGD-001 maksimum ağırlık 10 kg yerine 8 kg olursa ne olur?",
    )
    return simulation.simulate_change(
        report, request, selected_id="SGD-001",
        use_existing_rag=False, use_lm_studio=False,
    )


def _decide(package):
    accepted_test = False
    for proposal in package.proposals:
        if proposal.requirement_id == "SGD-001":
            package_logic.update_proposal(
                package, proposal.proposal_id, package_logic.DECISION_ACCEPT
            )
        elif proposal.category in {
            package_logic.CATEGORY_TEST,
            package_logic.CATEGORY_VERIFICATION,
            package_logic.CATEGORY_NEW_TEST,
        } and not accepted_test:
            package_logic.update_proposal(
                package, proposal.proposal_id, package_logic.DECISION_ACCEPT
            )
            accepted_test = True
        elif proposal.requirement_id == "STT-001":
            package_logic.update_proposal(
                package, proposal.proposal_id, package_logic.DECISION_REJECT
            )
        else:
            package_logic.update_proposal(
                package, proposal.proposal_id, package_logic.DECISION_DEFER
            )
    package_logic.mark_explicit_approval(package, "Test Onay Kurulu")


class SafeChangePackageTests(unittest.TestCase):
    def test_end_to_end_versioning_traceability_closure_pdf_and_excel(self):
        flat = _flat_data()
        original = deepcopy(flat)
        before = _report(flat)
        result = _simulation(before)
        package = package_logic.build_change_package(result, before)
        _decide(package)

        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "özgün_kaynak.txt"
            source.write_text("Özgün kaynak belge", encoding="utf-8")

            def validator(new_flat, new_hardware, stage):
                after = traceability.build_traceability_map(
                    "Güvenli Güncelleme Test Projesi",
                    flat_data=new_flat,
                    hardware_data=new_hardware,
                    document_sections=SECTIONS,
                    persist=False,
                    check_lm_studio=False,
                )
                rerun = simulation.simulate_change(
                    after,
                    result.change_request,
                    selected_id="SGD-001",
                    use_existing_rag=False,
                    use_lm_studio=False,
                ).to_dict()
                closure = package_logic.compare_closure(package, after, rerun)
                reports = stage / "reports"
                pdf = reporting.export_change_package_pdf(
                    reports / "rapor.pdf", package,
                    before_traceability=before,
                    after_traceability=after,
                    closure_summary=closure,
                )
                excel = reporting.export_change_package_excel(
                    reports / "rapor.xlsx", package,
                    before_traceability=before,
                    after_traceability=after,
                    closure_summary=closure,
                )
                return {
                    "post_traceability": after,
                    "post_simulation": rerun,
                    "closure_summary": closure,
                    "report_paths": {"pdf": pdf, "excel": excel},
                }

            applied = package_logic.apply_approved_changes(
                package,
                flat,
                source_paths=(source,),
                output_root=root / "change_control",
                validator=validator,
            )

            self.assertEqual(flat, original, "Özgün yapılandırılmış veri değişmemeli")
            self.assertEqual(applied.previous_version, 1)
            self.assertEqual(applied.new_version, 2)
            self.assertEqual(applied.new_flat_data["SGD-001"]["content"], "Sistemin maksimum ağırlığı 8 kg değerini aşmamalıdır.")
            self.assertEqual(applied.new_flat_data["STT-001"]["content"], original["STT-001"]["content"])
            self.assertTrue(Path(applied.backup_directory, "structured_data.v0001.json").is_file())
            self.assertTrue(any(Path(applied.backup_directory, "source_files").iterdir()))
            self.assertTrue(Path(applied.version_directory, "traceability.after.json").is_file())
            self.assertTrue(Path(applied.version_directory, "change_record.json").is_file())
            self.assertGreater(applied.closure_summary["resolved_count"], 0)
            self.assertGreater(applied.closure_summary["continuing_count"], 0)

            for document in applied.created_documents:
                path = Path(document)
                self.assertTrue(path.is_file())
                if path.suffix == ".pdf":
                    self.assertGreaterEqual(len(PdfReader(str(path)).pages), 1)

            pdf_path = Path(applied.report_paths["pdf"])
            excel_path = Path(applied.report_paths["excel"])
            pdf = PdfReader(str(pdf_path))
            self.assertGreaterEqual(len(pdf.pages), 3)
            pdf_text = "\n".join(page.extract_text() or "" for page in pdf.pages)
            self.assertIn("Güvenli Değişiklik Paketi", pdf_text)
            self.assertIn("Değişiklik Öncesi ve Sonrası", pdf_text)
            workbook = load_workbook(excel_path, data_only=False)
            self.assertEqual(tuple(workbook.sheetnames), reporting.SHEET_NAMES)
            self.assertEqual(workbook["Onay Kaydı"]["A1"].value, "Değişiklik Onay Kaydı")
            self.assertGreater(workbook["Değişiklik Listesi"].max_row, 5)
            workbook.close()

            stale_package = package_logic.build_change_package(
                _simulation(before), before
            )
            _decide(stale_package)
            with self.assertRaisesRegex(
                package_logic.ChangePackageError, "son yayımlanan sürümle uyuşmuyor"
            ):
                package_logic.apply_approved_changes(
                    stale_package,
                    flat,
                    output_root=root / "change_control",
                )
            current = json.loads(
                Path(applied.version_directory).parents[1]
                .joinpath("current.json")
                .read_text(encoding="utf-8")
            )
            self.assertEqual(current["version"], 2)

    def test_no_explicit_approval_and_stale_baseline_are_rejected_without_version(self):
        flat = _flat_data()
        report = _report(flat)
        package = package_logic.build_change_package(_simulation(report), report)
        main = next(item for item in package.proposals if item.requirement_id == "SGD-001")
        package_logic.update_proposal(package, main.proposal_id, package_logic.DECISION_ACCEPT)
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            with self.assertRaises(package_logic.ChangePackageError):
                package_logic.apply_approved_changes(package, flat, output_root=root)
            self.assertFalse(any(root.rglob("COMMIT.json")))

            package_logic.mark_explicit_approval(package, "Test Onayı")
            stale = deepcopy(flat)
            stale["SGD-001"]["content"] = "Bu içerik paket oluşturulduktan sonra değişti."
            with self.assertRaises(package_logic.ChangePackageError):
                package_logic.apply_approved_changes(package, stale, output_root=root)
            self.assertFalse(any(root.rglob("COMMIT.json")))

    def test_new_requirement_uses_user_id_and_creates_new_system_document_version(self):
        flat = _flat_data()
        report = _report(flat)
        request = simulation.ChangeRequest(
            requirement_id="SYS-REQ-777",
            current_value=None,
            proposed_value="Sistem teşhis kaydını en az 30 gün saklamalıdır.",
            reason="Bakım izlenebilirliği",
            requested_by="Müşteri",
            change_type=simulation.CHANGE_REQUIREMENT_ADD,
            assumptions=(),
            query="teşhis kaydı saklama gereksinimi",
        )
        simulated = simulation.simulate_change(
            report, request, use_existing_rag=False, use_lm_studio=False
        )
        package = package_logic.build_change_package(simulated, report)
        proposal = next(item for item in package.proposals if item.requirement_id == "SYS-REQ-777")
        self.assertEqual(proposal.target_kind, "new_requirement")
        self.assertEqual(proposal.document_type, "SGD")
        package_logic.update_proposal(package, proposal.proposal_id, package_logic.DECISION_ACCEPT)
        package_logic.mark_explicit_approval(package, "Müşteri Temsilcisi")
        with tempfile.TemporaryDirectory() as temp:
            applied = package_logic.apply_approved_changes(
                package, flat, output_root=Path(temp)
            )
            self.assertIn("SYS-REQ-777", applied.added_item_ids)
            self.assertEqual(applied.new_flat_data["SYS-REQ-777"]["type"], "SGD")
            self.assertNotIn("SYS-REQ-777", flat)

    def test_part_alternative_versions_hardware_pdf_docx_and_json(self):
        flat = _flat_data()
        hardware = {
            "HW-001": {
                "ID": "HW-001",
                "name": "Mevcut tahrik motoru",
                "description": "Mevcut tahrik motoru",
                "linked_requirements": ["STT-001"],
                "status": "approved",
            }
        }
        original_hardware = deepcopy(hardware)
        report = traceability.build_traceability_map(
            "Donanım Alternatifi Test Projesi",
            flat_data=flat,
            hardware_data=hardware,
            document_sections=SECTIONS,
            persist=False,
            check_lm_studio=False,
        )
        request = simulation.ChangeRequest(
            requirement_id="HW-001",
            current_value="Mevcut tahrik motoru",
            proposed_value="Düşük kütleli alternatif tahrik motoru",
            reason="Kütle hedefini sağlamak",
            requested_by="Mekanik Tasarım",
            change_type=simulation.CHANGE_PART_ALTERNATIVE,
            assumptions=(),
        )
        simulated = simulation.simulate_change(
            report, request, selected_id="HW-001",
            use_existing_rag=False, use_lm_studio=False,
        )
        package = package_logic.build_change_package(simulated, report)
        proposal = next(item for item in package.proposals if item.requirement_id == "HW-001")
        self.assertEqual(proposal.target_kind, "hardware_record")
        package_logic.update_proposal(package, proposal.proposal_id, package_logic.DECISION_ACCEPT)
        package_logic.mark_explicit_approval(package, "Donanım Değişiklik Kurulu")
        with tempfile.TemporaryDirectory() as temp:
            applied = package_logic.apply_approved_changes(
                package, flat, hardware_data=hardware, output_root=Path(temp)
            )
            self.assertEqual(hardware, original_hardware)
            self.assertEqual(
                applied.new_hardware_data["HW-001"]["description"],
                "Düşük kütleli alternatif tahrik motoru",
            )
            suffixes = {Path(path).suffix for path in applied.created_documents}
            self.assertTrue({".pdf", ".docx", ".json"}.issubset(suffixes))


if __name__ == "__main__":
    unittest.main()
