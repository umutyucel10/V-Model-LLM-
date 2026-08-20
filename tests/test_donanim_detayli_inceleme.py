# -*- coding: utf-8 -*-

from copy import deepcopy
from pathlib import Path
import tempfile
import unittest

from openpyxl import load_workbook
from pypdf import PdfReader

from donanim_detayli_inceleme import (
    alternative_comparison_rows, alternative_ids, breadcrumb, child_items,
    connection_rows, history_rows, overview, requirement_rows, state_rows,
    technical_rows,
)
from donanim_detayli_inceleme_raporlama import export_hardware_excel, export_hardware_pdf
from donanim_kartlari_model import MISSING_VALUE
import donanim_kartlari_yonetim as management


class HardwareDetailedReviewTests(unittest.TestCase):
    def setUp(self):
        self.catalog = management.sample_catalog()
        self.item = next(item for item in self.catalog["hardware_items"] if item["hardware_id"] == "SAMPLE-DCDC")
        self.trace = {
            "nodes": [
                {"id": "DEMO-REQ-PWR-001", "title": "28 V giriş gereksinimi", "description": "Dönüştürücü 28 V besleme ile çalışmalıdır.", "node_type": "Sistem gereksinimi", "v_model_level": "Sistem", "source_document": "SGD.pdf", "confidence_level": "Kesin"},
                {"id": "DEMO-TEST-PWR-001", "title": "Giriş gerilimi testi", "node_type": "Sistem doğrulama testi", "source_document": "STP.pdf"},
            ],
            "edges": [{"source_id": "DEMO-REQ-PWR-001", "target_id": "DEMO-TEST-PWR-001", "relationship": "verified_by"}],
        }

    def test_breadcrumb_and_children_use_real_parent_relations(self):
        self.assertEqual(
            breadcrumb(self.catalog, "SAMPLE-DCDC"),
            "Kontrol Sistemi  ›  Güç Alt Sistemi  ›  Güç Dağıtım Kartı  ›  DC/DC Dönüştürücü",
        )
        self.assertEqual(
            {item["hardware_id"] for item in child_items(self.catalog, "SAMPLE-PDB")},
            {"SAMPLE-DCDC", "SAMPLE-CUR"},
        )

    def test_technical_rows_never_turn_missing_values_into_zero(self):
        rows = {row["field"]: row for row in technical_rows(self.item)}
        self.assertEqual(rows["weight"]["value"], "18")
        self.assertEqual(rows["diameter"]["value"], MISSING_VALUE)
        self.assertNotEqual(rows["diameter"]["value"], "0")

    def test_requirements_include_linked_verification_test(self):
        rows = requirement_rows(self.item, self.trace)
        self.assertEqual(rows[0]["id"], "DEMO-REQ-PWR-001")
        self.assertIn("DEMO-TEST-PWR-001", rows[0]["tests"])
        self.assertEqual(rows[0]["source"], "SGD.pdf")

    def test_connection_and_state_views_are_available_with_sparse_data(self):
        connections = connection_rows(self.catalog, self.item, self.trace)
        self.assertTrue(any(row["type"] == "Üst sistem" for row in connections))
        states = state_rows(self.item)
        self.assertTrue(any(row["state"] == "Normal" for row in states))
        self.assertTrue(all(row["parameters"] == MISSING_VALUE for row in states))

    def test_alternative_comparison_and_single_alternative_payload(self):
        self.assertEqual(alternative_ids(self.catalog, "SAMPLE-DCDC"), ["SAMPLE-DCDC-B", "SAMPLE-DCDC-C"])
        comparison = alternative_comparison_rows(self.catalog, "SAMPLE-DCDC", "SAMPLE-DCDC-B")
        self.assertTrue(any(row["parameter"] == "Güç tüketimi" and row["assessment"] == "Olumlu" for row in comparison))
        payload = management.build_impact_payload(self.catalog, "SAMPLE-DCDC", "SAMPLE-DCDC-B")
        self.assertEqual(payload["alternatives"], ["DC/DC Alternatif B"])
        self.assertEqual(payload["hardware_context"]["alternative_ids"], ["SAMPLE-DCDC-B"])

    def test_manual_edit_is_recorded_in_change_history(self):
        overrides = management.empty_overrides("Örnek", self.catalog)
        management.set_field_override(overrides, "SAMPLE-DCDC", "manufacturer", "Yeni Üretici", self.catalog)
        rows = history_rows(overrides, "SAMPLE-DCDC")
        self.assertEqual(rows[0]["action"], "Manuel alan düzenlemesi")
        self.assertEqual(rows[0]["new_value"], "Yeni Üretici")

    def test_overview_marks_missing_engineering_data_explicitly(self):
        values = overview(self.item, self.catalog, self.trace)
        self.assertEqual(values["risks"], MISSING_VALUE)
        self.assertIn("Çalışma sıcaklığı", values["critical_limits"])

    def test_pdf_and_excel_reports_open_and_contain_expected_sections(self):
        overrides = management.empty_overrides("Örnek", self.catalog)
        management.set_field_override(overrides, "SAMPLE-DCDC", "manufacturer", "Rapor Üreticisi", self.catalog)
        with tempfile.TemporaryDirectory() as directory:
            pdf_path = export_hardware_pdf(Path(directory) / "kart.pdf", self.catalog, "SAMPLE-DCDC", self.trace, overrides)
            excel_path = export_hardware_excel(Path(directory) / "kart.xlsx", self.catalog, "SAMPLE-DCDC", self.trace, overrides)
            self.assertTrue(pdf_path.read_bytes().startswith(b"%PDF"))
            self.assertGreater(pdf_path.stat().st_size, 3000)
            reader = PdfReader(pdf_path)
            self.assertGreaterEqual(len(reader.pages), 2)
            extracted = "\n".join(page.extract_text() or "" for page in reader.pages)
            self.assertIn("DC/DC Dönüştürücü", extracted)
            self.assertIn("Teknik Özellikler", extracted)
            workbook = load_workbook(excel_path, read_only=False)
            self.assertEqual(
                workbook.sheetnames,
                ["Kimlik", "Teknik Özellikler", "Gereksinimler", "Bağlantılar", "Çalışma Durumları", "Alternatifler", "Kaynaklar", "Değişiklik Geçmişi"],
            )
            self.assertEqual(workbook["Kimlik"]["A1"].value, "Alan")
            self.assertEqual(workbook["Teknik Özellikler"].freeze_panes, "A2")
            workbook.close()


if __name__ == "__main__":
    unittest.main()
