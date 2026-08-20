# -*- coding: utf-8 -*-

import csv
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

import hardware_export_logic as exporter
import hardware_list_logic as hardware
from hardware_review_ui import parse_specifications_text


class HardwareReviewAndExportTests(unittest.TestCase):
    def setUp(self):
        self.flat_data = {
            "SGD-001": {
                "ID": "SGD-001",
                "type": "SGD",
                "content": "Sistem 28 VDC besleme hattından çalışmalıdır.",
            },
            "STT-001": {
                "ID": "STT-001",
                "type": "STT",
                "content": "Alt sistem -40°C ile +70°C arasında çalışmalıdır.",
            },
        }
        self.complete_record = {
            "ID": "HW-001",
            "category": "Güç Birimi",
            "description": "28 VDC güç dönüştürme birimi",
            "quantity": 1,
            "specifications": {
                "Giriş Gerilimi": "28 VDC",
                "Çalışma Sıcaklığı": "-40°C / +70°C",
            },
            "linked_requirements": ["SGD-001", "STT-001"],
            "status": "İnceleniyor",
            "risk": "Düşük",
            "manufacturer": "Örnek Üretici",
            "part_number": "PSU-28-001",
            "rationale": "SGD-001 ana besleme tahsisi için seçildi.",
            "review_note": "Katalog sayfası incelendi.",
        }

    def test_edit_returns_an_approved_record_to_review(self):
        registry = {"HW-001": {**self.complete_record, "status": "Onaylandı"}}

        updated = hardware.update_hardware_record(
            registry,
            "HW-001",
            {"quantity": 2, "review_note": "Adet güncellendi."},
        )

        self.assertEqual(updated["status"], "İnceleniyor")
        self.assertEqual(updated["quantity"], 2)
        self.assertEqual(updated["review_note"], "Adet güncellendi.")

    def test_approval_is_blocked_until_trace_catalog_and_specs_are_complete(self):
        registry = {
            "HW-001": {
                "ID": "HW-001",
                "description": "Güç birimi",
                "linked_requirements": ["SGD-001"],
                "status": "İnceleniyor",
            }
        }

        with self.assertRaisesRegex(ValueError, "Onay engelleri"):
            hardware.transition_hardware_status(
                registry,
                "HW-001",
                "Onaylandı",
                known_requirement_ids=self.flat_data,
            )

        registry["HW-001"] = dict(self.complete_record)
        approved = hardware.transition_hardware_status(
            registry,
            "HW-001",
            "Onaylandı",
            known_requirement_ids=self.flat_data,
        )
        self.assertEqual(approved["status"], "Onaylandı")

    def test_rejection_requires_and_keeps_engineer_reason(self):
        registry = {"HW-001": dict(self.complete_record)}

        with self.assertRaisesRegex(ValueError, "gerekçesi zorunludur"):
            hardware.transition_hardware_status(registry, "HW-001", "Reddedildi")

        rejected = hardware.transition_hardware_status(
            registry,
            "HW-001",
            "Reddedildi",
            review_note="Isıl marj yetersiz.",
        )
        self.assertEqual(rejected["status"], "Reddedildi")
        self.assertEqual(rejected["review_note"], "Isıl marj yetersiz.")

    def test_compatibility_report_detects_cross_record_conflicts_and_coverage_gap(self):
        first = dict(self.complete_record)
        first["linked_requirements"] = ["SGD-001"]
        second = {
            **self.complete_record,
            "ID": "HW-002",
            "manufacturer": "Farklı Üretici",
            "specifications": {"Giriş Gerilimi": "24 VDC"},
            "linked_requirements": ["SGD-001"],
        }

        issues = hardware.hardware_compatibility_report(
            {"HW-001": first, "HW-002": second}, self.flat_data
        )
        codes = {issue["code"] for issue in issues}

        self.assertIn("PART_MANUFACTURER_CONFLICT", codes)
        self.assertIn("DUPLICATE_ITEM", codes)
        self.assertIn("UNCOVERED_REQUIREMENT", codes)

    def test_specification_editor_parser_accepts_colon_and_equals(self):
        parsed = parse_specifications_text(
            "Giriş Gerilimi: 28 VDC\nÇıkış Gücü = DSB\n"
        )

        self.assertEqual(parsed, {
            "Giriş Gerilimi": "28 VDC",
            "Çıkış Gücü": "DSB",
        })
        with self.assertRaisesRegex(ValueError, "satırı 1"):
            parse_specifications_text("Biçimsiz satır")

    def test_excel_export_contains_bom_trace_specs_issues_and_formula_summary(self):
        registry = {"HW-001": dict(self.complete_record)}
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "bom.xlsx"
            result = exporter.export_hardware_excel(
                path, registry, self.flat_data, "Test Projesi"
            )
            workbook = load_workbook(path, data_only=False)

            self.assertEqual(result["record_count"], 1)
            self.assertEqual(
                workbook.sheetnames,
                ["Özet", "Donanım BOM", "Teknik Özellikler", "İzlenebilirlik", "Uyumluluk"],
            )
            self.assertTrue(str(workbook["Özet"]["B8"].value).startswith("=COUNTA"))
            self.assertEqual(workbook["Donanım BOM"]["A2"].value, "HW-001")
            self.assertEqual(workbook["Teknik Özellikler"].max_row, 3)
            self.assertEqual(workbook["İzlenebilirlik"].max_row, 3)
            self.assertEqual(workbook["Donanım BOM"].freeze_panes, "E2")

    def test_doors_export_uses_hardware_objects_and_trace_attributes(self):
        registry = {"HW-001": dict(self.complete_record)}
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "bom.csv"
            exporter.export_hardware_doors_csv(
                path, registry, self.flat_data, "Test Projesi"
            )
            with path.open("r", encoding="utf-8-sig", newline="") as stream:
                rows = list(csv.DictReader(stream))

            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["Object Identifier"], "HW-001")
            self.assertEqual(rows[0]["Object Type"], "Hardware BOM Item")
            self.assertEqual(rows[0]["Linked Requirements"], "SGD-001; STT-001")
            self.assertIn("Giriş Gerilimi=28 VDC", rows[0]["Specifications"])


if __name__ == "__main__":
    unittest.main()
