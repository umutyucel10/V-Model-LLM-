# -*- coding: utf-8 -*-

import copy
import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

import etki_analizi_logic as logic
import etki_analizi_raporlama as reporting


def sample_result():
    return logic.calculate_impact_analysis({
        "analysis_name": "İşlemci kartı değişikliği",
        "current_state": "Mevcut Kart A - sürüm 2",
        "change_reason": "Tedarik riski ve ömür devri maliyetini azaltma",
        "alternatives": ["Kart B", "Kart C", "Kart D"],
        "parameters": [
            {
                "name": "Performans",
                "current_value": 60,
                "alternative_values": {
                    "Kart B": 85,
                    "Kart C": 110,
                    "Kart D": "",
                },
                "unit": "puan",
                "weight": 50,
                "direction": "Yüksek daha iyi",
                "minimum": 0,
                "maximum": 100,
                "mandatory": True,
            },
            {
                "name": "Maliyet",
                "current_value": 100,
                "alternative_values": {
                    "Kart B": 80,
                    "Kart C": 60,
                    "Kart D": 75,
                },
                "unit": "bin TL",
                "weight": 30,
                "direction": "Düşük daha iyi",
                "minimum": 50,
                "maximum": 150,
                "mandatory": False,
            },
            {
                "name": "Çalışma sıcaklığı",
                "current_value": 65,
                "alternative_values": {
                    "Kart B": 60,
                    "Kart C": 75,
                    "Kart D": 70,
                },
                "unit": "°C",
                "weight": 20,
                "direction": "Düşük daha iyi",
                "minimum": 40,
                "maximum": 80,
                "mandatory": True,
            },
        ],
    })


class EtkiAnaliziRaporlamaTests(unittest.TestCase):
    def setUp(self):
        self.result = sample_result()
        self.fixed_time = datetime(2026, 7, 31, 10, 30)

    def test_pdf_is_readable_and_preserves_turkish_report_content(self):
        before = copy.deepcopy(self.result)
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "etki_analizi.pdf"
            export = reporting.export_impact_analysis_pdf(
                path,
                self.result,
                report_datetime=self.fixed_time,
            )

            reader = PdfReader(path)
            self.assertGreaterEqual(len(reader.pages), 2)
            self.assertGreater(
                float(reader.pages[0].mediabox.width),
                float(reader.pages[0].mediabox.height),
            )
            self.assertIn(
                "İşlemci kartı değişikliği",
                str(reader.metadata.title),
            )
            text = "\n".join(page.extract_text() or "" for page in reader.pages)

            self.assertIn("ETKİ ANALİZİ RAPORU", text)
            self.assertIn("Değişiklik nedeni", text)
            self.assertIn("Çalışma sıcaklığı", text)
            self.assertIn("Önerilen Alternatif", text)
            self.assertIn("Kart B", text)
            self.assertIn("Sayfa 1", text)
            self.assertEqual(export["format"], "pdf")
            self.assertEqual(Path(export["path"]), path.resolve())
            self.assertEqual(self.result, before)

    def test_excel_has_required_sheets_headers_values_and_formatting(self):
        before = copy.deepcopy(self.result)
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "etki_analizi.xlsx"
            export = reporting.export_impact_analysis_excel(
                path,
                self.result,
                report_datetime=self.fixed_time,
            )
            workbook = load_workbook(path, data_only=False)

            self.assertEqual(workbook.sheetnames, [
                "Özet",
                "Parametreler",
                "Fark Analizi",
                "Kriter Sonuçları",
                "Hesaplama",
            ])
            summary = workbook["Özet"]
            self.assertEqual(summary["A1"].value, "ETKİ ANALİZİ · SONUÇ ÖZETİ")
            self.assertEqual(summary["B3"].value, "İşlemci kartı değişikliği")
            self.assertEqual(summary["B4"].value, "Mevcut Kart A - sürüm 2")
            self.assertEqual(summary["B6"].value, self.fixed_time)
            self.assertEqual(summary["A10"].value, "Kart B")
            self.assertEqual(summary["B10"].value, 73.5)
            self.assertEqual(summary["C10"].value, "Uygun")
            self.assertEqual(summary.freeze_panes, "A10")
            self.assertEqual(len(summary.tables), 1)
            self.assertGreater(len(summary.conditional_formatting), 0)
            self.assertEqual(summary.row_dimensions[5].height, 42)
            self.assertEqual(
                next(iter(summary.tables.values())).autoFilter.ref, "A9:D12"
            )

            parameters = workbook["Parametreler"]
            self.assertEqual(parameters["A1"].value, "Alternatif")
            self.assertEqual(parameters["B1"].value, "Parametre Adı")
            self.assertEqual(parameters["C2"].value, 60)
            self.assertEqual(parameters["D2"].value, 85)
            self.assertEqual(parameters["F2"].value, 50)
            self.assertEqual(parameters["J2"].value, "Evet")
            self.assertEqual(parameters["K2"].value, 85)
            self.assertEqual(parameters.freeze_panes, "A2")
            self.assertEqual(len(parameters.tables), 1)
            self.assertEqual(parameters["F2"].number_format, '0.00"%"')
            self.assertEqual(
                next(iter(parameters.tables.values())).autoFilter.ref, "A1:L10"
            )

            differences = workbook["Fark Analizi"]
            self.assertEqual(differences["E2"].value, 25)
            self.assertAlmostEqual(differences["F2"].value, 41.6667, places=4)
            self.assertEqual(differences["F2"].number_format, '0.00"%"')
            self.assertEqual(differences["I2"].value, "Olumlu")
            self.assertEqual(differences.freeze_panes, "A2")

            criteria = workbook["Kriter Sonuçları"]
            self.assertEqual(criteria["C2"].value, "Evet")
            self.assertEqual(criteria["F2"].value, "Uygun")
            self.assertIn("Zorunlu kriter", criteria["G2"].value)

            calculation = workbook["Hesaplama"]
            self.assertEqual(calculation["A12"].value, "Parametre")
            self.assertEqual(calculation["B13"].value, 50)
            self.assertEqual(calculation["B13"].number_format, '0.00"%"')
            self.assertTrue(calculation.freeze_panes)
            self.assertEqual(export["sheet_count"], 5)
            self.assertEqual(Path(export["path"]), path.resolve())
            self.assertEqual(self.result, before)

    def test_invalid_result_is_rejected_without_creating_file(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "gecersiz.xlsx"
            with self.assertRaisesRegex(ValueError, "eksik alanlar"):
                reporting.export_impact_analysis_excel(
                    path,
                    {"analysis_name": "Eksik"},
                )
            self.assertFalse(path.exists())

    def test_reporting_does_not_change_calculation_results(self):
        fresh = sample_result()
        self.assertEqual(fresh["best_alternative"], {
            "alternative_name": "Kart B",
            "total_score": 73.5,
        })
        self.assertEqual(
            [item["status"] for item in fresh["alternatives"]],
            ["Uygun", "Uygun değil", "Veri eksik"],
        )


if __name__ == "__main__":
    unittest.main()

