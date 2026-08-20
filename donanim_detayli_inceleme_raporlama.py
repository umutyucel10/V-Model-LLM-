# -*- coding: utf-8 -*-
"""Donanım Detaylı İnceleme ekranının PDF kartı ve Excel çalışma kitabı."""

from __future__ import annotations

from datetime import datetime
from html import escape
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Image as ReportImage
from reportlab.platypus import LongTable, PageBreak, Paragraph, SimpleDocTemplate, Spacer, TableStyle

from donanim_detayli_inceleme import (
    alternative_comparison_rows, alternative_ids, breadcrumb, connection_rows,
    display, history_rows, overview, requirement_rows, source_rows, state_rows,
    technical_rows, gallery_entries,
)
from donanim_kartlari_model import MISSING_VALUE, clean_text
from hardware_image_generation import AI_CONCEPT_WARNING


BLUE = "0052CC"
NAVY = "17365D"
GRAPHITE = "3F4852"
LIGHT = "F3F5F7"
BORDER = "D8DEE5"
WHITE = "FFFFFF"


def _font_names() -> tuple[str, str]:
    regular, bold = "HardwareDetail", "HardwareDetail-Bold"
    if regular in pdfmetrics.getRegisteredFontNames():
        return regular, bold
    candidates = [
        ("/System/Library/Fonts/Supplemental/Arial.ttf", "/System/Library/Fonts/Supplemental/Arial Bold.ttf"),
        ("C:/Windows/Fonts/arial.ttf", "C:/Windows/Fonts/arialbd.ttf"),
    ]
    try:
        import reportlab
        root = Path(reportlab.__file__).resolve().parent / "fonts"
        candidates.append((str(root / "Vera.ttf"), str(root / "VeraBd.ttf")))
    except Exception:
        pass
    for regular_path, bold_path in candidates:
        if Path(regular_path).is_file() and Path(bold_path).is_file():
            pdfmetrics.registerFont(TTFont(regular, regular_path))
            pdfmetrics.registerFont(TTFont(bold, bold_path))
            return regular, bold
    return "Helvetica", "Helvetica-Bold"


def _p(value: Any, style: ParagraphStyle) -> Paragraph:
    return Paragraph(escape(display(value)).replace("\n", "<br/>"), style)


def _pdf_table(headers: Sequence[str], rows: Sequence[Sequence[Any]], widths: Sequence[float], styles: tuple[ParagraphStyle, ParagraphStyle]) -> LongTable:
    head, body = styles
    data = [[_p(value, head) for value in headers]] + [[_p(value, body) for value in row] for row in rows]
    table = LongTable(data, colWidths=list(widths), repeatRows=1, splitByRow=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{NAVY}")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), .3, colors.HexColor(f"#{BORDER}")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor(f"#{LIGHT}")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    return table


def _item(catalog: Mapping[str, Any], hardware_id: str) -> dict[str, Any]:
    for candidate in catalog.get("hardware_items", []):
        if isinstance(candidate, Mapping) and clean_text(candidate.get("hardware_id")) == hardware_id:
            return dict(candidate)
    raise ValueError("Raporlanacak donanım kartı bulunamadı.")


def export_hardware_pdf(
    path: str | Path, catalog: Mapping[str, Any], hardware_id: str,
    report: Mapping[str, Any] | None = None, overrides: Mapping[str, Any] | None = None,
) -> Path:
    item = _item(catalog, hardware_id)
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    regular, bold = _font_names()
    title = ParagraphStyle("H1", fontName=bold, fontSize=18, leading=22, textColor=colors.HexColor(f"#{NAVY}"), spaceAfter=7)
    section = ParagraphStyle("H2", fontName=bold, fontSize=11, leading=14, textColor=colors.HexColor(f"#{BLUE}"), spaceBefore=8, spaceAfter=5)
    head = ParagraphStyle("TH", fontName=bold, fontSize=7.2, leading=9, textColor=colors.white)
    body = ParagraphStyle("TD", fontName=regular, fontSize=7.2, leading=9.2, textColor=colors.HexColor(f"#{GRAPHITE}"))
    warning = ParagraphStyle(
        "AIWarning", fontName=bold, fontSize=8, leading=10,
        textColor=colors.HexColor("#8A5A00"), backColor=colors.HexColor("#FFF2CC"),
        borderPadding=6, spaceBefore=4, spaceAfter=6,
    )
    doc = SimpleDocTemplate(str(target), pagesize=landscape(A4), leftMargin=11 * mm, rightMargin=11 * mm, topMargin=15 * mm, bottomMargin=15 * mm)
    created = datetime.now().astimezone().strftime("%d.%m.%Y %H:%M")
    analysis = overview(item, catalog, report)

    def page(canvas, document):
        canvas.saveState(); canvas.setFont(regular, 7); canvas.setFillColor(colors.HexColor(f"#{GRAPHITE}"))
        canvas.drawString(11 * mm, 8 * mm, f"EHSİM · Donanım Kartı · {display(item.get('part_name'))}")
        canvas.drawRightString(landscape(A4)[0] - 11 * mm, 8 * mm, f"{created} · Sayfa {document.page}")
        canvas.restoreState()

    story = [_p("DONANIM DETAYLI İNCELEME", title), _p(f"{display(item.get('part_name'))} · {display(item.get('part_number'))}", section)]
    gallery = gallery_entries(item)
    cover = next((entry for entry in gallery if entry.get("is_cover")), gallery[0] if gallery else None)
    if cover and Path(str(cover.get("path"))).is_file():
        try:
            picture = ReportImage(str(cover["path"])); picture._restrictSize(92 * mm, 54 * mm)
            story.extend((picture, Spacer(1, 2 * mm)))
            if cover.get("is_ai"):
                story.append(_p(AI_CONCEPT_WARNING, warning))
                story.append(_pdf_table(
                    ["Görsel kaynağı", "Sağlayıcı", "Model", "Seed", "Kart sürümü"],
                    [[cover.get("source_type"), cover.get("provider"), cover.get("model"), cover.get("seed"), cover.get("source_card_version")]],
                    [45*mm, 45*mm, 55*mm, 25*mm, 45*mm], (head, body),
                ))
        except Exception:
            # Bozuk/uyumsuz görsel raporun tamamını engellemez.
            story.append(_p("Kapak görseli rapora eklenemedi; kart verileri korunarak raporlandı.", warning))
    identity = [
        ["Parça kimliği", hardware_id, "Üretici", display(item.get("manufacturer")), "Model", display(item.get("model_series"))],
        ["Tür", display(item.get("hardware_type")), "Yaşam döngüsü", display(item.get("lifecycle_status")), "Güven", display(item.get("confidence_score"))],
        ["Ürün ağacı", breadcrumb(catalog, hardware_id), "Sürüm", display(item.get("version")), "Güncelleme", display(item.get("updated_at"))],
    ]
    story.append(_pdf_table(["Alan", "Değer", "Alan", "Değer", "Alan", "Değer"], identity, [27*mm, 47*mm, 27*mm, 47*mm, 27*mm, 82*mm], (head, body)))
    parent_path = breadcrumb(catalog, hardware_id)
    parent_label = parent_path.rsplit("→", 1)[0].strip() if "→" in parent_path else MISSING_VALUE
    story += [
        _p("İzlenebilirlik Şeridi", section),
        _pdf_table(
            ["Üst Sistem", "Parça", "Gereksinim", "Test", "Alternatif"],
            [[
                parent_label, display(item.get("part_name")),
                f"{len(item.get('requirement_ids') or [])} bağlı gereksinim",
                f"{len(item.get('test_ids') or [])} bağlı test",
                f"{len(alternative_ids(catalog, hardware_id))} alternatif",
            ]],
            [55*mm, 55*mm, 50*mm, 45*mm, 50*mm], (head, body),
        ),
    ]
    story += [_p("Genel Bakış", section), _pdf_table(["Başlık", "İçerik"], [[key.replace("_", " ").title(), value] for key, value in analysis.items()], [47*mm, 210*mm], (head, body))]

    tech = technical_rows(item)
    story += [PageBreak(), _p("Teknik Özellikler", section), _pdf_table(
        ["Kategori", "Parametre", "Değer", "Birim", "Min", "Maks", "Tolerans", "Kaynak", "Konum", "Güven / Tür"],
        [[r["category"], r["parameter"], r["value"], r["unit"], r["minimum"], r["maximum"], r["tolerance"], r["source_document"], r["location"], f"{r['confidence']} / {r['certainty']}"] for r in tech],
        [20*mm, 31*mm, 27*mm, 13*mm, 18*mm, 18*mm, 20*mm, 34*mm, 31*mm, 38*mm], (head, body))]

    reqs = requirement_rows(item, report)
    story += [_p("Gereksinimler ve Testler", section), _pdf_table(
        ["Kimlik", "Gereksinim", "Seviye", "İlişki", "Durum", "Test", "Sonuç", "Kaynak", "Güven"],
        [[r[k] for k in ("id", "text", "level", "relation", "compliance", "tests", "test_result", "source", "confidence")] for r in reqs] or [[MISSING_VALUE]*9],
        [24*mm, 68*mm, 22*mm, 22*mm, 26*mm, 29*mm, 22*mm, 29*mm, 20*mm], (head, body))]

    connections = connection_rows(catalog, item, report)
    story += [_p("Sistem Bağlantıları", section), _pdf_table(["Yön", "Tür", "Kimlik", "Bağlantı", "Kaynak"], [[r[k] for k in ("direction", "type", "id", "name", "source")] for r in connections] or [[MISSING_VALUE]*5], [13*mm, 44*mm, 37*mm, 100*mm, 62*mm], (head, body))]
    story += [_p("Çalışma Durumları", section), _pdf_table(["Durum", "Değişen değerler", "Gereksinimler", "Parçalar", "Risk", "Test", "Beklenen davranış"], [[r[k] for k in ("state", "parameters", "requirements", "parts", "risks", "tests", "behavior")] for r in state_rows(item)], [27*mm, 42*mm, 36*mm, 36*mm, 35*mm, 35*mm, 45*mm], (head, body))]

    alternatives = alternative_ids(catalog, hardware_id)
    for alternative_id in alternatives:
        alternative = _item(catalog, alternative_id)
        story += [PageBreak(), _p(f"Alternatif Karşılaştırması · {display(alternative.get('part_name'))}", section), _pdf_table(
            ["Parametre", "Mevcut", "Alternatif", "Birim", "Değerlendirme"],
            [[r[k] for k in ("parameter", "current", "alternative", "unit", "assessment")] for r in alternative_comparison_rows(catalog, hardware_id, alternative_id)],
            [72*mm, 50*mm, 50*mm, 25*mm, 59*mm], (head, body))]

    sources = source_rows(item)
    story += [_p("Kaynak ve Kanıtlar", section), _pdf_table(["Alan", "Belge", "Konum", "Kanıt", "Yöntem", "Güven", "Tür"], [[r[k] for k in ("field", "document", "location", "evidence", "method", "confidence", "certainty")] for r in sources] or [[MISSING_VALUE]*7], [29*mm, 42*mm, 29*mm, 80*mm, 29*mm, 20*mm, 27*mm], (head, body))]
    history = history_rows(overrides, hardware_id)
    story += [_p("Değişiklik Geçmişi", section), _pdf_table(["Tarih", "İşlem", "Alan", "Önce", "Sonra", "Kullanıcı"], [[display(r.get(k)) for k in ("timestamp", "action", "field", "old_value", "new_value", "actor")] for r in history] or [[MISSING_VALUE]*6], [36*mm, 45*mm, 41*mm, 48*mm, 48*mm, 38*mm], (head, body)), Spacer(1, 4*mm)]
    doc.build(story, onFirstPage=page, onLaterPages=page)
    return target


def _sheet(workbook: Workbook, title: str, headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> None:
    ws = workbook.create_sheet(title=title)
    ws.append(list(headers))
    for row in rows:
        ws.append([display(value) for value in row])
    fill = PatternFill("solid", fgColor=NAVY); white = Font(color=WHITE, bold=True)
    edge = Side(style="thin", color=BORDER)
    for cell in ws[1]:
        cell.fill = fill; cell.font = white; cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(bottom=edge); cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    for index, column in enumerate(ws.columns, 1):
        width = max((len(str(cell.value or "")) for cell in column), default=8)
        ws.column_dimensions[get_column_letter(index)].width = min(max(width + 2, 12), 55)


def export_hardware_excel(
    path: str | Path, catalog: Mapping[str, Any], hardware_id: str,
    report: Mapping[str, Any] | None = None, overrides: Mapping[str, Any] | None = None,
) -> Path:
    item = _item(catalog, hardware_id); workbook = Workbook(); workbook.remove(workbook.active)
    identity = [("Parça adı", item.get("part_name")), ("Parça numarası", item.get("part_number")), ("Üretici", item.get("manufacturer")), ("Model", item.get("model_series")), ("Tür", item.get("hardware_type")), ("Sistem görevi", item.get("system_role")), ("Yaşam döngüsü", item.get("lifecycle_status")), ("Çalışma durumu", item.get("working_states")), ("Güven skoru", item.get("confidence_score")), ("Ürün ağacı", breadcrumb(catalog, hardware_id)), ("Sürüm", item.get("version")), ("Son güncelleme", item.get("updated_at"))]
    _sheet(workbook, "Kimlik", ["Alan", "Değer"], identity)
    tech = technical_rows(item); _sheet(workbook, "Teknik Özellikler", ["Kategori", "Parametre", "Değer", "Birim", "Minimum", "Maksimum", "Tolerans", "Durum Değeri", "Kaynak", "Bölüm", "Güven", "Bilgi Türü"], [[r[k] for k in ("category", "parameter", "value", "unit", "minimum", "maximum", "tolerance", "state_value", "source_document", "location", "confidence", "certainty")] for r in tech])
    req = requirement_rows(item, report); _sheet(workbook, "Gereksinimler", ["Kimlik", "Metin", "Seviye", "İlişki", "Karşılama", "Test", "Test Sonucu", "Kaynak", "Güven"], [[r[k] for k in ("id", "text", "level", "relation", "compliance", "tests", "test_result", "source", "confidence")] for r in req])
    con = connection_rows(catalog, item, report); _sheet(workbook, "Bağlantılar", ["Yön", "Tür", "Kimlik", "Ad", "Kaynak"], [[r[k] for k in ("direction", "type", "id", "name", "source")] for r in con])
    states = state_rows(item); _sheet(workbook, "Çalışma Durumları", ["Durum", "Değişen Değerler", "Gereksinimler", "Parçalar", "Riskler", "Testler", "Beklenen Davranış"], [[r[k] for k in ("state", "parameters", "requirements", "parts", "risks", "tests", "behavior")] for r in states])
    alt_rows = []
    for alt_id in alternative_ids(catalog, hardware_id):
        alt_name = _item(catalog, alt_id).get("part_name")
        alt_rows.extend([[alt_name, r["parameter"], r["current"], r["alternative"], r["unit"], r["assessment"]] for r in alternative_comparison_rows(catalog, hardware_id, alt_id)])
    _sheet(workbook, "Alternatifler", ["Alternatif", "Parametre", "Mevcut", "Alternatif", "Birim", "Değerlendirme"], alt_rows)
    sources = source_rows(item); _sheet(workbook, "Kaynaklar", ["Alan", "Belge", "Konum", "Kanıt", "Yöntem", "Güven", "Bilgi Türü", "Dosya"], [[r[k] for k in ("field", "document", "location", "evidence", "method", "confidence", "certainty", "path")] for r in sources])
    history = history_rows(overrides, hardware_id); _sheet(workbook, "Değişiklik Geçmişi", ["Tarih", "İşlem", "Alan", "Önce", "Sonra", "Kullanıcı"], [[r.get(k) for k in ("timestamp", "action", "field", "old_value", "new_value", "actor")] for r in history])
    target = Path(path); target.parent.mkdir(parents=True, exist_ok=True); workbook.save(target); return target


__all__ = ["export_hardware_excel", "export_hardware_pdf"]
