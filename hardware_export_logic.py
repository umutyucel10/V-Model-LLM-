# -*- coding: utf-8 -*-
"""Akıllı Donanım Listesi için Excel ve IBM DOORS dışa aktarımı."""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

import hardware_list_logic


NAVY = "17365D"
GRAPHITE = "3F4852"
LIGHT_BLUE = "DCE6F1"
LIGHT_GRAY = "EEF1F4"
BORDER = "D8DEE5"
SUCCESS = "D9EAD3"
WARNING = "FFF2CC"
DANGER = "F4CCCC"
MUTED = "E7E6E6"
WHITE = "FFFFFF"

BOM_HEADERS = (
    "ID",
    "Durum",
    "Risk",
    "Kategori",
    "Donanım Tanımı",
    "Adet",
    "Üretici",
    "Parça Numarası",
    "Bağlı Gereksinimler",
    "Öneri Güveni",
    "Gerekçe",
    "Mühendis Notu",
    "Kaynak Alıntı",
    "Uyumluluk Sonucu",
)


def _clean(value: Any) -> str:
    return " ".join(str(value or "").split())


def _selected_records(
    registry: Mapping[str, Mapping[str, Any]],
    approved_only: bool,
) -> list[dict[str, Any]]:
    records = hardware_list_logic.exportable_hardware_records(
        registry, approved_only=approved_only
    )
    if not records:
        scope = "onaylanmış" if approved_only else ""
        raise ValueError(f"Dışa aktarılacak {scope} donanım kaydı bulunamadı.".strip())
    return records


def _blocking_export_issues(
    issues: Iterable[Mapping[str, Any]],
    selected_ids: set[str],
) -> list[Mapping[str, Any]]:
    return [
        issue
        for issue in issues
        if issue.get("severity") == "Hata"
        and (
            issue.get("item_id") in selected_ids
            or issue.get("code") == "UNCOVERED_REQUIREMENT"
        )
    ]


def _prepare_export(
    registry: Mapping[str, Mapping[str, Any]],
    flat_data: Mapping[str, Mapping[str, Any]],
    approved_only: bool,
) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    records = _selected_records(registry, approved_only)
    issues = hardware_list_logic.hardware_compatibility_report(registry, flat_data)
    if approved_only:
        blockers = _blocking_export_issues(
            issues, {record["ID"] for record in records}
        )
        if blockers:
            preview = "\n• ".join(
                _clean(issue.get("message")) for issue in blockers[:6]
            )
            raise ValueError(
                "Onaylı BOM dışa aktarımı uyumluluk hataları nedeniyle engellendi:\n• "
                + preview
            )
    return records, issues


def _style_title(ws, cell_range: str, text: str) -> None:
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = text
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(name="Segoe UI", size=14, bold=True, color=WHITE)
    cell.alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[cell.row].height = 28


def _style_header(ws, row: int, start_col: int, end_col: int) -> None:
    thin = Side(style="thin", color=BORDER)
    for cell in ws.iter_cols(
        min_col=start_col, max_col=end_col, min_row=row, max_row=row
    ):
        item = cell[0]
        item.fill = PatternFill("solid", fgColor=GRAPHITE)
        item.font = Font(name="Segoe UI", size=9, bold=True, color=WHITE)
        item.alignment = Alignment(vertical="center", wrap_text=True)
        item.border = Border(bottom=thin)
    ws.row_dimensions[row].height = 28


def _style_body(ws, min_row: int, max_row: int, max_col: int) -> None:
    if max_row < min_row:
        return
    thin = Side(style="thin", color=BORDER)
    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=1, max_col=max_col
    ):
        for cell in row:
            cell.font = Font(name="Segoe UI", size=9, color="222222")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)


def _add_table(ws, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _issue_text_by_item(issues: Iterable[Mapping[str, Any]]) -> dict[str, str]:
    grouped: dict[str, list[str]] = {}
    for issue in issues:
        item_id = _clean(issue.get("item_id"))
        if not item_id:
            continue
        grouped.setdefault(item_id, []).append(
            f"{_clean(issue.get('severity'))}: {_clean(issue.get('message'))}"
        )
    return {key: " | ".join(values) for key, values in grouped.items()}


def export_hardware_excel(
    path: str | Path,
    registry: Mapping[str, Mapping[str, Any]],
    flat_data: Mapping[str, Mapping[str, Any]],
    project_name: str = "Proje",
    approved_only: bool = False,
) -> dict[str, Any]:
    """Çok sayfalı, izlenebilir ve düzenlenebilir BOM çalışma kitabı üretir."""
    records, issues = _prepare_export(registry, flat_data, approved_only)
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    issue_text = _issue_text_by_item(issues)

    workbook = Workbook()
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcOnSave = True
    summary_ws = workbook.active
    summary_ws.title = "Özet"
    bom_ws = workbook.create_sheet("Donanım BOM")
    specs_ws = workbook.create_sheet("Teknik Özellikler")
    trace_ws = workbook.create_sheet("İzlenebilirlik")
    issues_ws = workbook.create_sheet("Uyumluluk")

    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False

    _style_title(summary_ws, "A1:F1", "AKILLI DONANIM LİSTESİ · BOM ÖZETİ")
    summary_ws["A3"] = "Proje"
    summary_ws["B3"] = _clean(project_name) or "Proje"
    summary_ws["A4"] = "Kapsam"
    summary_ws["B4"] = "Yalnızca Onaylı" if approved_only else "Tüm Kayıtlar"
    summary_ws["A5"] = "Oluşturma Zamanı"
    summary_ws["B5"] = datetime.now().replace(microsecond=0)
    summary_ws["B5"].number_format = "yyyy-mm-dd hh:mm"
    for cell in ("A3", "A4", "A5"):
        summary_ws[cell].font = Font(name="Segoe UI", size=9, bold=True, color=GRAPHITE)
    summary_ws["A7"] = "Gösterge"
    summary_ws["B7"] = "Değer"
    summary_rows = (
        ("Toplam Donanım", f"=COUNTA('Donanım BOM'!A2:A{len(records) + 1})"),
        ("Onaylı", f'=COUNTIF(\'Donanım BOM\'!B2:B{len(records) + 1},"Onaylandı")'),
        ("İnceleniyor", f'=COUNTIF(\'Donanım BOM\'!B2:B{len(records) + 1},"İnceleniyor")'),
        ("Yüksek Risk", f'=COUNTIF(\'Donanım BOM\'!C2:C{len(records) + 1},"Yüksek")'),
        ("Uyumluluk Hatası", f'=COUNTIF(\'Uyumluluk\'!A2:A{max(2, len(issues) + 1)},"Hata")'),
        ("Uyumluluk Uyarısı", f'=COUNTIF(\'Uyumluluk\'!A2:A{max(2, len(issues) + 1)},"Uyarı")'),
    )
    for row_index, (label, formula) in enumerate(summary_rows, start=8):
        summary_ws.cell(row=row_index, column=1, value=label)
        summary_ws.cell(row=row_index, column=2, value=formula)
    _style_header(summary_ws, 7, 1, 2)
    _style_body(summary_ws, 8, 13, 2)
    summary_ws.column_dimensions["A"].width = 26
    summary_ws.column_dimensions["B"].width = 24

    bom_ws.append(BOM_HEADERS)
    for record in records:
        confidence = record.get("confidence")
        bom_ws.append([
            record["ID"], record["status"], record["risk"], record["category"],
            record["description"], record["quantity"], record["manufacturer"],
            record["part_number"], "; ".join(record["linked_requirements"]),
            confidence if confidence is not None else None,
            record["rationale"], record.get("review_note", ""),
            record["source_excerpt"], issue_text.get(record["ID"], "Uygun"),
        ])
    _style_header(bom_ws, 1, 1, len(BOM_HEADERS))
    _style_body(bom_ws, 2, len(records) + 1, len(BOM_HEADERS))
    _add_table(bom_ws, "HardwareBOMTable", f"A1:N{len(records) + 1}")
    bom_ws.freeze_panes = "E2"
    for row_index in range(2, len(records) + 2):
        bom_ws.row_dimensions[row_index].height = 42
    widths = (12, 15, 12, 20, 44, 8, 20, 20, 28, 14, 38, 32, 40, 48)
    for index, width in enumerate(widths, start=1):
        bom_ws.column_dimensions[chr(64 + index)].width = width
    for cell in bom_ws["J"][1:]:
        cell.number_format = "0%"
    status_validation = DataValidation(
        type="list", formula1='"Önerilen,İnceleniyor,Onaylandı,Reddedildi"'
    )
    risk_validation = DataValidation(
        type="list", formula1='"Belirsiz,Düşük,Orta,Yüksek"'
    )
    bom_ws.add_data_validation(status_validation)
    bom_ws.add_data_validation(risk_validation)
    status_validation.add(f"B2:B{len(records) + 1}")
    risk_validation.add(f"C2:C{len(records) + 1}")
    bom_ws.conditional_formatting.add(
        f"B2:B{len(records) + 1}",
        FormulaRule(formula=['$B2="Onaylandı"'], fill=PatternFill("solid", fgColor=SUCCESS)),
    )
    bom_ws.conditional_formatting.add(
        f"B2:B{len(records) + 1}",
        FormulaRule(formula=['$B2="Reddedildi"'], fill=PatternFill("solid", fgColor=MUTED)),
    )
    bom_ws.conditional_formatting.add(
        f"C2:C{len(records) + 1}",
        FormulaRule(formula=['$C2="Yüksek"'], fill=PatternFill("solid", fgColor=DANGER)),
    )
    bom_ws.conditional_formatting.add(
        f"N2:N{len(records) + 1}",
        FormulaRule(formula=['ISNUMBER(SEARCH("Hata:",$N2))'], fill=PatternFill("solid", fgColor=DANGER)),
    )

    specs_ws.append(("Donanım ID", "Teknik Özellik", "Değer"))
    spec_rows = []
    for record in records:
        for name, value in record["specifications"].items():
            spec_rows.append((record["ID"], name, value))
            specs_ws.append((record["ID"], name, value))
    _style_header(specs_ws, 1, 1, 3)
    _style_body(specs_ws, 2, len(spec_rows) + 1, 3)
    if spec_rows:
        _add_table(specs_ws, "HardwareSpecsTable", f"A1:C{len(spec_rows) + 1}")
    specs_ws.freeze_panes = "A2"
    for row_index in range(2, len(spec_rows) + 2):
        specs_ws.row_dimensions[row_index].height = 22
    for column, width in zip(("A", "B", "C"), (14, 32, 38)):
        specs_ws.column_dimensions[column].width = width

    requirement_map = {
        record["requirement_id"]: record
        for record in hardware_list_logic.eligible_requirement_records(flat_data)
    }
    trace_ws.append((
        "Gereksinim ID", "Gereksinim Türü", "Gereksinim Metni",
        "Donanım ID", "Donanım Tanımı", "Donanım Durumu",
    ))
    trace_rows = []
    for record in records:
        for requirement_id in record["linked_requirements"]:
            requirement = requirement_map.get(requirement_id, {})
            row = (
                requirement_id,
                requirement.get("requirement_type", "Bilinmiyor"),
                requirement.get("content", "Bilinmeyen gereksinim"),
                record["ID"], record["description"], record["status"],
            )
            trace_rows.append(row)
            trace_ws.append(row)
    _style_header(trace_ws, 1, 1, 6)
    _style_body(trace_ws, 2, len(trace_rows) + 1, 6)
    if trace_rows:
        _add_table(trace_ws, "HardwareTraceTable", f"A1:F{len(trace_rows) + 1}")
    trace_ws.freeze_panes = "D2"
    for row_index in range(2, len(trace_rows) + 2):
        trace_ws.row_dimensions[row_index].height = 38
    for column, width in zip(("A", "B", "C", "D", "E", "F"), (18, 16, 54, 14, 42, 16)):
        trace_ws.column_dimensions[column].width = width

    issues_ws.append(("Seviye", "Kayıt", "Kod", "Açıklama"))
    for issue in issues:
        issues_ws.append((
            issue["severity"], issue["item_id"], issue["code"], issue["message"]
        ))
    _style_header(issues_ws, 1, 1, 4)
    _style_body(issues_ws, 2, len(issues) + 1, 4)
    if issues:
        _add_table(issues_ws, "HardwareIssuesTable", f"A1:D{len(issues) + 1}")
    issues_ws.freeze_panes = "A2"
    for row_index in range(2, len(issues) + 2):
        issues_ws.row_dimensions[row_index].height = 34
    for column, width in zip(("A", "B", "C", "D"), (12, 16, 30, 70)):
        issues_ws.column_dimensions[column].width = width
    if issues:
        issues_ws.conditional_formatting.add(
            f"A2:A{len(issues) + 1}",
            FormulaRule(formula=['$A2="Hata"'], fill=PatternFill("solid", fgColor=DANGER)),
        )
        issues_ws.conditional_formatting.add(
            f"A2:A{len(issues) + 1}",
            FormulaRule(formula=['$A2="Uyarı"'], fill=PatternFill("solid", fgColor=WARNING)),
        )

    workbook.save(output_path)
    return {
        "path": str(output_path),
        "record_count": len(records),
        "issue_count": len(issues),
        "approved_only": approved_only,
    }


def export_hardware_doors_csv(
    path: str | Path,
    registry: Mapping[str, Mapping[str, Any]],
    flat_data: Mapping[str, Mapping[str, Any]],
    project_name: str = "Proje",
    approved_only: bool = False,
) -> dict[str, Any]:
    """Donanım nesnelerini DOORS'a aktarılabilir UTF-8 CSV olarak üretir."""
    records, issues = _prepare_export(registry, flat_data, approved_only)
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    issue_text = _issue_text_by_item(issues)
    headers = (
        "Object Identifier", "Object Heading", "Object Text", "Object Type",
        "Status", "Risk", "Quantity", "Manufacturer", "Part Number",
        "Linked Requirements", "Specifications", "Rationale",
        "Review Note", "Compatibility", "Project",
    )
    with output_path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.writer(stream, quoting=csv.QUOTE_ALL, lineterminator="\r\n")
        writer.writerow(headers)
        for record in records:
            specifications = "; ".join(
                f"{name}={value}" for name, value in record["specifications"].items()
            )
            writer.writerow((
                record["ID"], record["category"], record["description"],
                "Hardware BOM Item", record["status"], record["risk"],
                record["quantity"], record["manufacturer"], record["part_number"],
                "; ".join(record["linked_requirements"]), specifications,
                record["rationale"], record.get("review_note", ""),
                issue_text.get(record["ID"], "Uygun"), _clean(project_name) or "Proje",
            ))
    return {
        "path": str(output_path),
        "record_count": len(records),
        "issue_count": len(issues),
        "approved_only": approved_only,
    }
