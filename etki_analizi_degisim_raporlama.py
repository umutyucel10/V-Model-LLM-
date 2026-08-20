# -*- coding: utf-8 -*-
"""Güvenli gereksinim değişiklik paketini PDF ve Excel olarak raporlar.

Bu modül hesaplama veya belge güncellemesi yapmaz. Yalnızca simülasyonda
hesaplanmış sonuçları, kullanıcı kararlarını ve sürümleme/kapanış kayıtlarını
değiştirmeden sunar.
"""

from __future__ import annotations

from datetime import datetime
from html import escape
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table as ExcelTable
from openpyxl.worksheet.table import TableStyleInfo
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    HRFlowable,
    KeepTogether,
    LongTable,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from etki_analizi_degisim_paketi import (
    ChangePackage,
    DECISION_ACCEPT,
    DECISION_DEFER,
    DECISION_EDIT,
    DECISION_REJECT,
    DECISIONS,
)
from etki_analizi_raporlama import _register_pdf_fonts


NAVY = "17365D"
GRAPHITE = "3F4852"
PAPER = "FFFFFF"
LIGHT_BLUE = "DCE6F1"
LIGHT_GRAY = "EEF1F4"
BORDER = "D8DEE5"
SUCCESS = "D9EAD3"
SUCCESS_TEXT = "217A43"
WARNING = "FFF2CC"
WARNING_TEXT = "9A6400"
DANGER = "F4CCCC"
DANGER_TEXT = "B42318"
MISSING = "E7E6E6"
MISSING_TEXT = "5C666D"
ORANGE = "FCE4D6"

SHEET_NAMES = (
    "Etki Analizi",
    "İzlenebilirlik",
    "Değişiklik Listesi",
    "Etkilenen Belgeler",
    "Test Planı",
    "Risk ve Aksiyon",
    "Mühendislik Fikirleri",
    "Onay Kaydı",
)


def _clean(value: Any) -> str:
    return " ".join(str(value or "").split())


def _mapping(value: Any) -> dict[str, Any]:
    if isinstance(value, Mapping):
        return dict(value)
    if hasattr(value, "to_dict"):
        return dict(value.to_dict())
    return {}


def _package(value: ChangePackage | Mapping[str, Any]) -> ChangePackage:
    package = value if isinstance(value, ChangePackage) else ChangePackage.from_mapping(value)
    if not package.change_id or not package.project_id:
        raise ValueError("Raporlanacak değişiklik paketi geçerli değil.")
    return package


def _list(value: Any) -> list[Any]:
    return list(value) if isinstance(value, Sequence) and not isinstance(value, (str, bytes)) else []


def _status_fill(status: Any) -> str:
    text = _clean(status)
    if text in {DECISION_ACCEPT, "Çözüldü", "Uygun", "Kesin", "Düşük"}:
        return SUCCESS
    if text in {DECISION_REJECT, "Kritik", "Uygun değil"}:
        return DANGER
    if text in {DECISION_DEFER, DECISION_EDIT, "Yüksek", "Orta", "Uyarı"}:
        return WARNING if text != "Yüksek" else ORANGE
    return MISSING


def _severity_order(value: Any) -> int:
    return {"Kritik": 4, "Yüksek": 3, "Orta": 2, "Düşük": 1}.get(_clean(value), 0)


def _pdf_text(value: Any) -> str:
    text = str(value if value not in (None, "") else "-")
    # ReportLab'in birlikte gelen Vera yazı tipi ok karakterini taşımıyor.
    # İz yolunu kare glif yerine platformlar arası güvenli ASCII ile göster.
    text = text.replace("→", "->").replace("⇒", "=>")
    return escape(text).replace("\n", "<br/>")


def _pdf_table(
    rows: Sequence[Sequence[Any]],
    widths: Sequence[float],
    body: ParagraphStyle,
    header: ParagraphStyle,
    statuses: Sequence[Any] | None = None,
) -> LongTable:
    data = []
    for row_index, row in enumerate(rows):
        style = header if row_index == 0 else body
        data.append([Paragraph(_pdf_text(value), style) for value in row])
    table = LongTable(data, colWidths=list(widths), repeatRows=1, splitByRow=1, hAlign="LEFT")
    commands: list[tuple] = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{NAVY}")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor(f"#{BORDER}")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    for row_index in range(1, len(data)):
        fill = LIGHT_GRAY if row_index % 2 == 0 else PAPER
        if statuses and row_index - 1 < len(statuses):
            fill = _status_fill(statuses[row_index - 1])
        commands.append(("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor(f"#{fill}")))
    table.setStyle(TableStyle(commands))
    return table


def _impact_rows(package: ChangePackage) -> list[dict[str, Any]]:
    simulation = package.simulation_snapshot
    return [dict(item) for item in simulation.get("impacts", []) if isinstance(item, Mapping)]


def _edge_rows(traceability: Mapping[str, Any]) -> list[dict[str, Any]]:
    return [dict(item) for item in traceability.get("edges", []) if isinstance(item, Mapping)]


def _closure_counts(closure: Mapping[str, Any]) -> tuple[str, str, str]:
    if not closure:
        return "-", "-", "-"
    return (
        str(closure.get("resolved_count", 0)),
        str(closure.get("continuing_count", 0)),
        str(closure.get("new_conflict_count", 0)),
    )


def export_change_package_pdf(
    path: str | Path,
    package: ChangePackage | Mapping[str, Any],
    *,
    before_traceability: Mapping[str, Any] | None = None,
    after_traceability: Mapping[str, Any] | None = None,
    closure_summary: Mapping[str, Any] | None = None,
) -> Path:
    """Sekiz rapor grubunu içeren denetlenebilir PDF değişiklik raporu üretir."""
    package = _package(package)
    before = dict(before_traceability or package.baseline_traceability)
    after = dict(after_traceability or {})
    closure = dict(closure_summary or {})
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    regular, bold = _register_pdf_fonts()
    report_time = datetime.now().astimezone()

    title_style = ParagraphStyle("ChangeTitle", fontName=bold, fontSize=18, leading=21, textColor=colors.HexColor(f"#{NAVY}"))
    section_style = ParagraphStyle("ChangeSection", fontName=bold, fontSize=12, leading=15, textColor=colors.HexColor(f"#{NAVY}"), spaceBefore=6, spaceAfter=5)
    body = ParagraphStyle("ChangeBody", fontName=regular, fontSize=7.2, leading=9.2, textColor=colors.HexColor(f"#{GRAPHITE}"))
    head = ParagraphStyle("ChangeHead", fontName=bold, fontSize=7, leading=8.5, textColor=colors.white)
    small = ParagraphStyle("ChangeSmall", fontName=regular, fontSize=6.3, leading=8, textColor=colors.HexColor(f"#{GRAPHITE}"))

    def footer(canvas: Any, doc: Any) -> None:
        canvas.saveState()
        canvas.setFont(regular, 6.5)
        canvas.setFillColor(colors.HexColor(f"#{MISSING_TEXT}"))
        canvas.drawString(12 * mm, 8 * mm, f"{package.project_name} · {package.change_id} · {report_time.strftime('%d.%m.%Y %H:%M')}")
        canvas.drawRightString(landscape(A4)[0] - 12 * mm, 8 * mm, f"Sayfa {doc.page}")
        canvas.restoreState()

    document = SimpleDocTemplate(
        str(output), pagesize=landscape(A4), leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=12 * mm, bottomMargin=13 * mm,
        title=f"Etki Analizi ve Değişiklik Paketi {package.change_id}",
        author="Etki Analizi Uygulaması",
    )
    story: list[Any] = [
        Paragraph("Etki Analizi ve Güvenli Değişiklik Paketi", title_style),
        Paragraph(_pdf_text(f"{package.project_name} · {package.change_id} · Durum: {package.status}"), body),
        Spacer(1, 3 * mm),
        HRFlowable(width="100%", thickness=1, color=colors.HexColor(f"#{NAVY}")),
        Spacer(1, 3 * mm),
    ]

    accepted = [item for item in package.proposals if item.decision == DECISION_ACCEPT]
    rejected = [item for item in package.proposals if item.decision == DECISION_REJECT]
    deferred = [item for item in package.proposals if item.decision in {DECISION_DEFER, DECISION_EDIT}]
    impacts = _impact_rows(package)
    critical = sum(_clean(item.get("impact_level")) == "Kritik" for item in impacts)
    resolved, continuing, new_conflicts = _closure_counts(closure)
    story.extend([
        Paragraph("1. Yönetici Özeti", section_style),
        _pdf_table([
            ["Alan", "Değer", "Alan", "Değer"],
            ["Değişiklik özeti", package.request_summary, "Değişiklik isteyen", package.change_request.get("requested_by")],
            ["Etkilenen öğe", len(impacts), "Kritik etki", critical],
            ["Etkilenen belge", len(package.affected_documents), "Öneri", len(package.proposals)],
            ["Kabul / Ret / Bekleyen", f"{len(accepted)} / {len(rejected)} / {len(deferred)}", "Onaylayan", package.approval_actor or "Açık onay verilmedi"],
            ["Çözülen / Devam eden", f"{resolved} / {continuing}", "Yeni çelişki", new_conflicts],
        ], [39 * mm, 94 * mm, 39 * mm, 94 * mm], body, head),
        Spacer(1, 4 * mm),
    ])

    request = package.change_request
    story.extend([
        Paragraph("2. Değişiklik Öncesi ve Sonrası", section_style),
        _pdf_table([
            ["Gereksinim", "Mevcut içerik/değer", "Önerilen içerik/değer", "Tür", "Neden"],
            [request.get("requirement_id") or _clean((package.selected_item or {}).get("id")), request.get("current_value"), request.get("proposed_value"), request.get("change_type"), request.get("reason")],
        ], [33 * mm, 68 * mm, 68 * mm, 39 * mm, 58 * mm], body, head),
    ])

    left_types = {"Müşteri/paydaş gereksinimi", "Sistem gereksinimi", "Alt sistem gereksinimi", "Fonksiyon", "Tasarım kararı", "Parça/bileşen", "Mekanik arayüz", "Elektriksel arayüz", "Yazılımsal arayüz"}
    right_types = {"Doğrulama kriteri", "Birim testi", "Entegrasyon testi", "Sistem doğrulama testi", "Müşteri kabul/geçerleme testi"}
    for heading, allowed in (("3. V-Model Sol Kol Etkileri", left_types), ("4. V-Model Sağ Kol Etkileri", right_types)):
        selected = [item for item in impacts if _clean(item.get("node_type")) in allowed]
        rows = [["Kimlik", "Tür", "Etki", "Puan", "Güven", "Gerekçe", "İzlenebilirlik yolu"]]
        rows.extend([
            [item.get("item_id"), item.get("node_type"), item.get("impact_level"), item.get("impact_score"), item.get("confidence_level"), item.get("rationale"), _mapping(item.get("traceability_path")).get("display_path")]
            for item in selected
        ])
        story.extend([
            Paragraph(heading, section_style),
            _pdf_table(rows if len(rows) > 1 else [rows[0], ["-", "-", "-", "-", "-", "Bu kolda etki bulunmadı.", "-"]], [23 * mm, 34 * mm, 19 * mm, 15 * mm, 24 * mm, 76 * mm, 75 * mm], small, head, [item.get("impact_level") for item in selected]),
        ])

    story.append(PageBreak())
    story.append(Paragraph("5. Etki Yolları ve Kaynak Kanıtları", section_style))
    paths = _list(package.simulation_snapshot.get("impact_paths"))
    path_rows = [["Yol", "Sınıf", "Derinlik", "Güven", "İlişkiler"]]
    for item in paths:
        row = _mapping(item)
        path_rows.append([row.get("display_path"), row.get("classification"), row.get("depth"), row.get("confidence_level"), " → ".join(row.get("relationships") or [])])
    story.append(_pdf_table(path_rows if len(path_rows) > 1 else [path_rows[0], ["-", "-", "-", "-", "Etki yolu bulunmadı."]], [98 * mm, 39 * mm, 19 * mm, 32 * mm, 78 * mm], body, head))

    story.append(Paragraph("6. V-Model İzlenebilirlik Matrisi", section_style))
    node_map = {
        str(item.get("id")): item
        for item in (after or before).get("nodes", [])
        if isinstance(item, Mapping)
    }
    matrix_rows = [["Kaynak", "Kaynak türü", "İlişki", "Hedef", "Hedef türü", "Güven", "Kaynak belge", "Kanıt"]]
    for edge in _edge_rows(after or before):
        source_id = edge.get("source_id", edge.get("source"))
        target_id = edge.get("target_id", edge.get("target"))
        matrix_rows.append([
            source_id,
            _mapping(node_map.get(str(source_id))).get("node_type"),
            edge.get("relationship_type", edge.get("relationship")),
            target_id,
            _mapping(node_map.get(str(target_id))).get("node_type"),
            edge.get("confidence_level"),
            edge.get("source_document"),
            edge.get("evidence_text", edge.get("evidence")),
        ])
    story.append(_pdf_table(matrix_rows if len(matrix_rows) > 1 else [matrix_rows[0], ["-", "-", "-", "-", "-", "-", "-", "Bağlantı bulunmadı."]], [24 * mm, 31 * mm, 27 * mm, 24 * mm, 31 * mm, 23 * mm, 39 * mm, 67 * mm], small, head))

    story.append(Paragraph("7. Etkilenen Belge Listesi", section_style))
    document_rows = [["Belge", "Öneri sayısı", "Kategoriler", "Bölümler", "Sürümleme durumu"]]
    for item in package.affected_documents:
        versioning = "Yeni sürüm" if any(
            proposal.document_name == item.get("document_name")
            and proposal.decision == DECISION_ACCEPT
            for proposal in package.proposals
        ) else "Değişiklik yok"
        document_rows.append([
            item.get("document_name"), item.get("proposal_count"),
            ", ".join(item.get("categories") or []),
            ", ".join(item.get("sections") or []), versioning,
        ])
    story.append(_pdf_table(document_rows if len(document_rows) > 1 else [document_rows[0], ["-", "-", "-", "-", "Etkilenen belge bulunmadı."]], [59 * mm, 25 * mm, 67 * mm, 75 * mm, 40 * mm], body, head))

    story.append(Paragraph("8. Test Güncelleme Planı", section_style))
    test_categories = {"Test prosedürleri", "Yeni test önerileri", "Doğrulama kriterleri", "Kabul kriterleri"}
    test_proposals = [item for item in package.proposals if item.category in test_categories]
    test_rows = [["Karar", "Kategori", "Test/kriter", "Mevcut", "Güncelleme", "Etki yolu", "Risk"]]
    for item in test_proposals:
        test_rows.append([item.decision, item.category, item.requirement_id, item.current_text, item.proposed_text, item.impact_path, item.risk_level])
    story.append(_pdf_table(test_rows if len(test_rows) > 1 else [test_rows[0], ["-", "-", "-", "-", "Yeni test güncellemesi bulunmadı.", "-", "-"]], [20 * mm, 37 * mm, 29 * mm, 49 * mm, 54 * mm, 58 * mm, 19 * mm], small, head, [item.decision for item in test_proposals]))

    story.append(Paragraph("9. Risk ve Aksiyon Listesi", section_style))
    risk_rows = [["Kategori", "Seviye", "Olasılık", "Şiddet", "Risk puanı", "Gerekçe", "Etkilenen öğeler"]]
    for risk in sorted(package.risks, key=lambda item: _severity_order(item.get("impact_level")), reverse=True):
        risk_rows.append([risk.get("category"), risk.get("impact_level"), risk.get("probability"), risk.get("severity"), risk.get("risk_score"), risk.get("rationale"), ", ".join(risk.get("impacted_items") or [])])
    story.append(_pdf_table(risk_rows if len(risk_rows) > 1 else [risk_rows[0], ["-", "-", "-", "-", "-", "Risk bulunmadı.", "-"]], [34 * mm, 21 * mm, 18 * mm, 18 * mm, 20 * mm, 98 * mm, 57 * mm], body, head, [item.get("impact_level") for item in package.risks]))

    story.append(Paragraph("10. Mühendislik Fikirleri", section_style))
    idea_rows = [["Kategori", "Öneri", "Gerekçe", "Beklenen fayda", "Yeni risk", "Gerekli doğrulama", "Durum"]]
    for idea in package.engineering_ideas:
        idea_rows.append([idea.get("category"), idea.get("suggestion"), idea.get("rationale"), idea.get("expected_benefit"), idea.get("new_risk"), idea.get("required_verification"), idea.get("status")])
    story.append(_pdf_table(idea_rows if len(idea_rows) > 1 else [idea_rows[0], ["-", "-", "-", "-", "-", "-", "Fikir üretilmedi."]], [32 * mm, 48 * mm, 44 * mm, 38 * mm, 35 * mm, 42 * mm, 27 * mm], small, head))

    story.append(PageBreak())
    story.append(Paragraph("11. Gereksinim Değişiklik Listesi ve Onay Kaydı", section_style))
    proposal_rows = [["Karar", "Kategori", "Belge / bölüm", "Kimlik", "Mevcut", "Önerilen", "Gerekçe", "Etki yolu", "Risk"]]
    for item in package.proposals:
        proposal_rows.append([item.decision, item.category, f"{item.document_name}\n{item.section}", item.requirement_id, item.current_text, item.proposed_text, item.rationale, item.impact_path, item.risk_level])
    story.append(_pdf_table(proposal_rows, [17 * mm, 27 * mm, 32 * mm, 20 * mm, 37 * mm, 37 * mm, 40 * mm, 38 * mm, 18 * mm], small, head, [item.decision for item in package.proposals]))

    story.extend([
        Spacer(1, 5 * mm),
        Paragraph("Onaylanan ve reddedilen değişiklikler", section_style),
        _pdf_table([
            ["Kayıt", "Değer"],
            ["Açık uygulama onayı", "Evet" if package.approval_confirmed else "Hayır"],
            ["Onaylayan", package.approval_actor or "-"],
            ["Onay zamanı", package.approval_at or "-"],
            ["Onaylanan", ", ".join(item.requirement_id for item in accepted) or "-"],
            ["Reddedilen", ", ".join(item.requirement_id for item in rejected) or "-"],
            ["Ertelenen/düzenlenecek", ", ".join(item.requirement_id for item in deferred) or "-"],
        ], [55 * mm, 211 * mm], body, head),
        KeepTogether([
            Paragraph("Varsayımlar ve Açık Sorular", section_style),
            _pdf_table([
                ["Tür", "Kayıt"],
                *[["Varsayım", value] for value in package.assumptions],
                *[["Açık soru", value] for value in package.open_questions],
            ] or [["Tür", "Kayıt"], ["-", "Kayıt yok."]], [40 * mm, 226 * mm], body, head),
        ]),
    ])
    document.build(story, onFirstPage=footer, onLaterPages=footer)
    return output


THIN = Side(style="thin", color=BORDER)


def _excel_title(ws: Any, title: str, subtitle: str) -> int:
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    cell = ws.cell(1, 1, title)
    cell.font = Font(name="Segoe UI", size=16, bold=True, color=PAPER)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
    ws.cell(2, 1, subtitle).font = Font(name="Segoe UI", size=9, color=GRAPHITE)
    ws.cell(2, 1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    return 4


def _write_table(ws: Any, start_row: int, headers: Sequence[str], rows: Iterable[Sequence[Any]], name: str) -> tuple[int, int]:
    rows = list(rows)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, col, header)
        cell.font = Font(name="Segoe UI", size=9, bold=True, color=PAPER)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = Border(bottom=THIN)
    for row_index, values in enumerate(rows, start=start_row + 1):
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_index, col, value)
            cell.font = Font(name="Segoe UI", size=9, color=GRAPHITE)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(bottom=THIN)
            if row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    if rows:
        reference = f"A{start_row}:{get_column_letter(len(headers))}{start_row + len(rows)}"
        table = ExcelTable(displayName=name, ref=reference)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
        ws.add_table(table)
        ws.auto_filter.ref = reference
    return start_row, start_row + len(rows)


def _decision_colors(ws: Any, column: int, first: int, last: int) -> None:
    for row in range(first, last + 1):
        value = _clean(ws.cell(row, column).value)
        ws.cell(row, column).fill = PatternFill("solid", fgColor=_status_fill(value))
        ws.cell(row, column).font = Font(name="Segoe UI", size=9, bold=True, color={DECISION_ACCEPT: SUCCESS_TEXT, DECISION_REJECT: DANGER_TEXT}.get(value, WARNING_TEXT if value in {DECISION_DEFER, DECISION_EDIT} else MISSING_TEXT))


def _finish_sheet(ws: Any, freeze: str = "A5") -> None:
    ws.freeze_panes = freeze
    for column in range(1, ws.max_column + 1):
        values = [_clean(ws.cell(row, column).value) for row in range(1, min(ws.max_row, 250) + 1)]
        width = max([len(value) for value in values] + [8]) + 2
        ws.column_dimensions[get_column_letter(column)].width = min(max(width, 11), 48)
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.oddHeader.center.text = "&B" + ws.title
    ws.oddFooter.right.text = "Sayfa &P / &N"


def export_change_package_excel(
    path: str | Path,
    package: ChangePackage | Mapping[str, Any],
    *,
    before_traceability: Mapping[str, Any] | None = None,
    after_traceability: Mapping[str, Any] | None = None,
    closure_summary: Mapping[str, Any] | None = None,
) -> Path:
    """Sekiz veri grubunu ayrı çalışma sayfalarına yerleştirir."""
    package = _package(package)
    before = dict(before_traceability or package.baseline_traceability)
    after = dict(after_traceability or {})
    closure = dict(closure_summary or {})
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    subtitle = f"{package.project_name} · {package.change_id} · {package.status}"

    ws = workbook.create_sheet(SHEET_NAMES[0])
    row = _excel_title(ws, "Etki Analizi Raporu", subtitle)
    summary_rows = [
        ("Değişiklik özeti", package.request_summary),
        ("Gereksinim", package.change_request.get("requirement_id") or _clean((package.selected_item or {}).get("id"))),
        ("Mevcut değer", package.change_request.get("current_value")),
        ("Önerilen değer", package.change_request.get("proposed_value")),
        ("Değişiklik nedeni", package.change_request.get("reason")),
        ("İsteyen taraf", package.change_request.get("requested_by")),
        ("Etkilenen öğe", len(_impact_rows(package))),
        ("Kritik etki", sum(_clean(item.get("impact_level")) == "Kritik" for item in _impact_rows(package))),
        ("Etkilenen belge", len(package.affected_documents)),
        ("Onay durumu", "Açık onay verildi" if package.approval_confirmed else "Açık onay bekleniyor"),
    ]
    _, end = _write_table(ws, row, ("Alan", "Değer"), summary_rows, "T_EtkiOzeti")
    impact_rows = [[item.get("item_id"), item.get("node_type"), item.get("impact_level"), item.get("impact_score"), item.get("risk_score"), item.get("confidence_level"), item.get("rationale"), _mapping(item.get("traceability_path")).get("display_path")] for item in _impact_rows(package)]
    _write_table(ws, end + 3, ("Kimlik", "Tür", "Etki seviyesi", "Etki puanı", "Risk puanı", "Güven", "Gerekçe", "İzlenebilirlik yolu"), impact_rows, "T_EtkiOgeleri")
    _finish_sheet(ws)

    ws = workbook.create_sheet(SHEET_NAMES[1])
    row = _excel_title(ws, "V-Model İzlenebilirlik Matrisi", subtitle)
    nodes_before = {str(item.get("id")): item for item in before.get("nodes", []) if isinstance(item, Mapping)}
    nodes_after = {str(item.get("id")): item for item in after.get("nodes", []) if isinstance(item, Mapping)}
    trace_rows = []
    for edge in _edge_rows(after or before):
        source_id = edge.get("source_id", edge.get("source"))
        target_id = edge.get("target_id", edge.get("target"))
        relationship = edge.get("relationship_type", edge.get("relationship"))
        evidence = edge.get("evidence_text", edge.get("evidence"))
        source = nodes_after.get(str(source_id)) or nodes_before.get(str(source_id)) or {}
        target = nodes_after.get(str(target_id)) or nodes_before.get(str(target_id)) or {}
        trace_rows.append([source_id, source.get("node_type"), relationship, target_id, target.get("node_type"), edge.get("confidence_level"), edge.get("source_document"), evidence])
    _write_table(ws, row, ("Kaynak", "Kaynak türü", "İlişki", "Hedef", "Hedef türü", "Güven", "Belge", "Kanıt"), trace_rows, "T_Izlenebilirlik")
    _finish_sheet(ws)

    ws = workbook.create_sheet(SHEET_NAMES[2])
    row = _excel_title(ws, "Gereksinim Değişiklik Listesi", subtitle)
    proposal_rows = [[item.proposal_id, item.category, item.document_name, item.section, item.requirement_id, item.current_text, item.proposed_text, item.rationale, item.impact_path, item.risk_level, item.decision] for item in package.proposals]
    _, end = _write_table(ws, row, ("Öneri kimliği", "Kategori", "Belge", "Bölüm", "Gereksinim", "Mevcut metin", "Önerilen metin", "Gerekçe", "Etki yolu", "Risk", "Kullanıcı kararı"), proposal_rows, "T_Degisiklikler")
    if end >= row + 1:
        validation = DataValidation(type="list", formula1='"' + ",".join(DECISIONS) + '"', allow_blank=False)
        ws.add_data_validation(validation)
        validation.add(f"K{row + 1}:K{end}")
        _decision_colors(ws, 11, row + 1, end)
    _finish_sheet(ws)

    ws = workbook.create_sheet(SHEET_NAMES[3])
    row = _excel_title(ws, "Etkilenen Belge Listesi", subtitle)
    document_rows = [[item.get("document_name"), item.get("proposal_count"), ", ".join(item.get("categories") or []), ", ".join(item.get("sections") or []), "Yeni sürüm" if any(p.document_name == item.get("document_name") and p.decision == DECISION_ACCEPT for p in package.proposals) else "Değişiklik yok"] for item in package.affected_documents]
    _write_table(ws, row, ("Belge", "Öneri sayısı", "Kategoriler", "Bölümler", "Sürümleme durumu"), document_rows, "T_EtkilenenBelgeler")
    _finish_sheet(ws)

    ws = workbook.create_sheet(SHEET_NAMES[4])
    row = _excel_title(ws, "Test Güncelleme Planı", subtitle)
    test_rows = [[item.proposal_id, item.category, item.document_name, item.requirement_id, item.current_text, item.proposed_text, item.impact_path, item.risk_level, item.decision] for item in package.proposals if item.category in {"Test prosedürleri", "Yeni test önerileri", "Doğrulama kriterleri", "Kabul kriterleri"}]
    _, end = _write_table(ws, row, ("Öneri", "Kategori", "Belge", "Test/kriter", "Mevcut", "Güncelleme", "Etki yolu", "Risk", "Karar"), test_rows, "T_TestPlani")
    if end >= row + 1:
        _decision_colors(ws, 9, row + 1, end)
    _finish_sheet(ws)

    ws = workbook.create_sheet(SHEET_NAMES[5])
    row = _excel_title(ws, "Risk ve Aksiyon Listesi", subtitle)
    risk_rows = [[item.get("category"), item.get("impact_level"), item.get("probability"), item.get("severity"), item.get("risk_score"), item.get("confidence_level"), item.get("rationale"), ", ".join(item.get("impacted_items") or []), item.get("source_evidence")] for item in package.risks]
    _, risk_end = _write_table(ws, row, ("Kategori", "Seviye", "Olasılık", "Şiddet", "Risk puanı", "Güven", "Gerekçe", "Etkilenen öğeler", "Kaynak kanıt"), risk_rows, "T_Riskler")
    action_rows = [[item.proposal_id, item.proposed_text, item.impact_path, item.risk_level, item.decision] for item in package.proposals if item.category == "Risk azaltma faaliyetleri"]
    _, action_end = _write_table(ws, risk_end + 3, ("Aksiyon", "Faaliyet", "Etki yolu", "Risk", "Karar"), action_rows, "T_RiskAksiyonlari")
    if action_end > risk_end + 3:
        _decision_colors(ws, 5, risk_end + 4, action_end)
    _finish_sheet(ws)

    ws = workbook.create_sheet(SHEET_NAMES[6])
    row = _excel_title(ws, "Mühendislik Fikirleri", subtitle)
    idea_rows = [[item.get("suggestion_id"), item.get("category"), item.get("suggestion"), item.get("rationale"), item.get("expected_benefit"), item.get("new_risk"), ", ".join(item.get("affected_items") or []), item.get("required_verification"), item.get("source_or_assumption"), item.get("status")] for item in package.engineering_ideas]
    _write_table(ws, row, ("Kimlik", "Kategori", "Öneri", "Gerekçe", "Beklenen fayda", "Yeni risk", "Etkilenenler", "Gerekli doğrulama", "Kaynak/varsayım", "Durum"), idea_rows, "T_MuhendislikFikirleri")
    _finish_sheet(ws)

    ws = workbook.create_sheet(SHEET_NAMES[7])
    row = _excel_title(ws, "Değişiklik Onay Kaydı", subtitle)
    approval_rows = [
        ("Değişiklik kimliği", package.change_id),
        ("Paket durumu", package.status),
        ("Açık onay", "Evet" if package.approval_confirmed else "Hayır"),
        ("Onaylayan", package.approval_actor),
        ("Onay zamanı", package.approval_at),
        ("Onay özeti", package.approval_digest),
        ("Çözülen öneri", closure.get("resolved_count")),
        ("Devam eden öneri", closure.get("continuing_count")),
        ("Yeni çelişki", closure.get("new_conflict_count")),
        ("Yeniden simülasyon", closure.get("rerun_status")),
    ]
    _, end = _write_table(ws, row, ("Kayıt", "Değer"), approval_rows, "T_OnayOzeti")
    decision_rows = [[item.proposal_id, item.requirement_id, item.decision, item.rationale, item.current_text, item.proposed_text] for item in package.proposals]
    _, final_end = _write_table(ws, end + 3, ("Öneri", "Öğe", "Karar", "Gerekçe", "Önce", "Sonra"), decision_rows, "T_OnayKararlari")
    if final_end > end + 3:
        _decision_colors(ws, 3, end + 4, final_end)
    _finish_sheet(ws)

    workbook.properties.title = f"Etki Analizi {package.change_id}"
    workbook.properties.subject = "Güvenli gereksinim değişikliği ve onay kaydı"
    workbook.properties.creator = "Etki Analizi Uygulaması"
    workbook.save(output)
    return output


__all__ = ["SHEET_NAMES", "export_change_package_excel", "export_change_package_pdf"]
