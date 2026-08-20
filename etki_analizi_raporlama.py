# -*- coding: utf-8 -*-
"""Etki Analizi sonuçlarını PDF ve Excel olarak raporlar.

Bu modül hesaplama yapmaz; ``etki_analizi_logic.calculate_impact_analysis``
tarafından üretilmiş sonuç sözlüğünü değiştirmeden sunar.
"""

from __future__ import annotations

from datetime import datetime
from html import escape
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table as ExcelTable
from openpyxl.worksheet.table import TableStyleInfo
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    HRFlowable,
    KeepTogether,
    LongTable,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


STATUS_SUITABLE = "Uygun"
STATUS_UNSUITABLE = "Uygun değil"
STATUS_MISSING = "Veri eksik"
STATUS_WARNING = "Kabul sınırı dışında"
STATUS_MANDATORY_FAIL = "Zorunlu kriter sağlanmadı"

NAVY = "17365D"
GRAPHITE = "3F4852"
LIGHT_BLUE = "DCE6F1"
LIGHT_GRAY = "EEF1F4"
BORDER = "D8DEE5"
SUCCESS = "D9EAD3"
SUCCESS_TEXT = "217A43"
WARNING = "FFF2CC"
WARNING_TEXT = "9A6400"
DANGER = "F4CCCC"
DANGER_TEXT = "B42318"
MISSING = "E7E6E6"
MISSING_TEXT = "5C666D"
WHITE = "FFFFFF"


def _clean(value: Any) -> str:
    return " ".join(str(value or "").split())


def _validate_result(result: Mapping[str, Any]) -> None:
    if not isinstance(result, Mapping):
        raise ValueError("Raporlanacak Etki Analizi sonucu geçerli değil.")
    required = (
        "analysis_name",
        "current_state",
        "change_reason",
        "normalized_weights",
        "alternatives",
        "best_alternative",
        "calculation_explanation",
    )
    missing = [key for key in required if key not in result]
    if missing:
        raise ValueError(
            "Etki Analizi sonucu eksik alanlar içeriyor: "
            + ", ".join(missing)
            + "."
        )
    alternatives = result.get("alternatives")
    if not isinstance(alternatives, Sequence) or isinstance(
        alternatives, (str, bytes)
    ) or not alternatives:
        raise ValueError("Raporlanacak alternatif sonucu bulunamadı.")
    if not isinstance(result.get("normalized_weights"), Mapping):
        raise ValueError("Normalize ağırlık bilgisi geçerli değil.")


def _format_number(value: Any, suffix: str = "") -> str:
    if value is None:
        return "Veri eksik"
    number = float(value)
    text = f"{number:.2f}".rstrip("0").rstrip(".")
    return f"{text}{suffix}"


def _paragraph_text(value: Any) -> str:
    text = str(value if value is not None else "Veri eksik")
    return escape(text).replace("\n", "<br/>")


def _status_fill(status: str) -> str:
    if status in {STATUS_UNSUITABLE, STATUS_MANDATORY_FAIL}:
        return DANGER
    if status in {STATUS_WARNING, "Uyarı"}:
        return WARNING
    if status == STATUS_MISSING:
        return MISSING
    if status == STATUS_SUITABLE:
        return SUCCESS
    return LIGHT_GRAY


def _status_text_color(status: str) -> str:
    if status in {STATUS_UNSUITABLE, STATUS_MANDATORY_FAIL}:
        return DANGER_TEXT
    if status in {STATUS_WARNING, "Uyarı"}:
        return WARNING_TEXT
    if status == STATUS_MISSING:
        return MISSING_TEXT
    if status == STATUS_SUITABLE:
        return SUCCESS_TEXT
    return GRAPHITE


def _difference_evaluation(criterion: Mapping[str, Any]) -> str:
    difference = criterion.get("difference")
    if difference is None:
        return STATUS_MISSING
    if float(difference) == 0:
        return "Nötr"
    direction = _clean(criterion.get("direction")).casefold()
    higher_is_better = direction in {
        "yüksek daha iyi",
        "yuksek daha iyi",
        "higher is better",
        "high",
    }
    positive = float(difference) > 0 if higher_is_better else float(difference) < 0
    return "Olumlu" if positive else "Olumsuz"


def _criterion_explanation(criterion: Mapping[str, Any]) -> str:
    status = _clean(criterion.get("status"))
    if status == STATUS_MISSING:
        return "Mevcut değer, alternatif değer veya kabul sınırlarından en az biri eksik."
    if status == STATUS_MANDATORY_FAIL:
        return "Zorunlu kriterin alternatif değeri kabul sınırlarının dışında."
    if status == STATUS_WARNING:
        return "Zorunlu olmayan kriterin değeri kabul sınırlarının dışında."
    if criterion.get("mandatory"):
        return "Zorunlu kriter kabul sınırları içinde sağlandı."
    return "Kriter kabul sınırları içinde."


def _collect_warnings(result: Mapping[str, Any]) -> list[dict[str, str]]:
    warnings: list[dict[str, str]] = []
    for alternative in result["alternatives"]:
        alternative_name = _clean(alternative.get("alternative_name"))
        if alternative.get("has_missing_data"):
            warnings.append({
                "level": STATUS_MISSING,
                "alternative": alternative_name,
                "parameter": "Genel",
                "message": "Eksik veri nedeniyle toplam puan hesaplanamadı.",
            })
        for criterion in alternative.get("criteria", []):
            status = _clean(criterion.get("status"))
            if status == STATUS_MISSING:
                warnings.append({
                    "level": STATUS_MISSING,
                    "alternative": alternative_name,
                    "parameter": _clean(criterion.get("parameter_name")),
                    "message": _criterion_explanation(criterion),
                })
            elif status == STATUS_WARNING:
                warnings.append({
                    "level": "Uyarı",
                    "alternative": alternative_name,
                    "parameter": _clean(criterion.get("parameter_name")),
                    "message": _criterion_explanation(criterion),
                })
            elif status == STATUS_MANDATORY_FAIL:
                warnings.append({
                    "level": STATUS_UNSUITABLE,
                    "alternative": alternative_name,
                    "parameter": _clean(criterion.get("parameter_name")),
                    "message": _criterion_explanation(criterion),
                })
    if not result.get("best_alternative"):
        warnings.append({
            "level": "Uyarı",
            "alternative": "-",
            "parameter": "Öneri",
            "message": "Uygun ve eksiksiz bir alternatif belirlenemedi.",
        })
    return warnings


def _register_pdf_fonts() -> tuple[str, str]:
    regular_name = "ImpactReportFont"
    bold_name = "ImpactReportFont-Bold"
    registered = set(pdfmetrics.getRegisteredFontNames())
    if regular_name in registered and bold_name in registered:
        return regular_name, bold_name
    try:
        import reportlab

        fonts_dir = Path(reportlab.__file__).resolve().parent / "fonts"
        pdfmetrics.registerFont(TTFont(regular_name, str(fonts_dir / "Vera.ttf")))
        pdfmetrics.registerFont(TTFont(bold_name, str(fonts_dir / "VeraBd.ttf")))
        return regular_name, bold_name
    except Exception:
        return "Helvetica", "Helvetica-Bold"


def _pdf_table(
    rows: list[list[Any]],
    widths: list[float],
    body_style: ParagraphStyle,
    header_style: ParagraphStyle,
    row_statuses: Sequence[str] | None = None,
    font_name: str = "Helvetica",
) -> LongTable:
    data: list[list[Paragraph]] = []
    for row_index, row in enumerate(rows):
        style = header_style if row_index == 0 else body_style
        data.append([Paragraph(_paragraph_text(value), style) for value in row])
    table = LongTable(
        data,
        colWidths=widths,
        repeatRows=1,
        splitByRow=1,
        hAlign="LEFT",
    )
    commands: list[tuple] = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{NAVY}")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LINEBELOW", (0, 0), (-1, 0), 0.8, colors.HexColor(f"#{NAVY}")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor(f"#{BORDER}")),
        ("FONTNAME", (0, 1), (-1, -1), font_name),
    ]
    if row_statuses:
        for index, status in enumerate(row_statuses, start=1):
            commands.append((
                "BACKGROUND",
                (0, index),
                (-1, index),
                colors.HexColor(f"#{_status_fill(status)}"),
            ))
    else:
        commands.append((
            "ROWBACKGROUNDS",
            (0, 1),
            (-1, -1),
            [colors.white, colors.HexColor(f"#{LIGHT_GRAY}")],
        ))
    table.setStyle(TableStyle(commands))
    return table


def export_impact_analysis_pdf(
    path: str | Path,
    result: Mapping[str, Any],
    report_datetime: datetime | None = None,
) -> dict[str, Any]:
    """Mevcut Etki Analizi sonucundan sayfalı, Türkçe PDF raporu üretir."""
    _validate_result(result)
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    created_at = report_datetime or datetime.now()
    regular, bold = _register_pdf_fonts()

    title_style = ParagraphStyle(
        "ImpactTitle",
        fontName=bold,
        fontSize=19,
        leading=23,
        textColor=colors.HexColor(f"#{NAVY}"),
        spaceAfter=3,
    )
    subtitle_style = ParagraphStyle(
        "ImpactSubtitle",
        fontName=regular,
        fontSize=9,
        leading=13,
        textColor=colors.HexColor(f"#{MISSING_TEXT}"),
    )
    section_style = ParagraphStyle(
        "ImpactSection",
        fontName=bold,
        fontSize=11,
        leading=14,
        textColor=colors.HexColor(f"#{NAVY}"),
        spaceBefore=8,
        spaceAfter=5,
        keepWithNext=True,
    )
    body_style = ParagraphStyle(
        "ImpactBody",
        fontName=regular,
        fontSize=7.2,
        leading=9.4,
        textColor=colors.HexColor("#222222"),
        alignment=TA_LEFT,
    )
    small_style = ParagraphStyle(
        "ImpactSmall",
        parent=body_style,
        fontSize=6.4,
        leading=8.1,
    )
    header_style = ParagraphStyle(
        "ImpactTableHeader",
        fontName=bold,
        fontSize=6.7,
        leading=8.2,
        textColor=colors.white,
        alignment=TA_CENTER,
    )
    callout_style = ParagraphStyle(
        "ImpactCallout",
        fontName=bold,
        fontSize=10,
        leading=14,
        textColor=colors.HexColor(f"#{SUCCESS_TEXT}"),
    )

    story: list[Any] = [
        Paragraph("ETKİ ANALİZİ RAPORU", title_style),
        Paragraph(
            f"Rapor tarihi: {created_at.strftime('%d.%m.%Y %H:%M')}",
            subtitle_style,
        ),
        Spacer(1, 3 * mm),
        HRFlowable(
            width="100%",
            thickness=1.2,
            color=colors.HexColor(f"#{NAVY}"),
        ),
        Spacer(1, 4 * mm),
        Paragraph("1. Analiz Bilgileri", section_style),
    ]
    metadata = [
        ["Alan", "Değer"],
        ["Analiz adı", result["analysis_name"]],
        ["Mevcut parça veya durum", result["current_state"]],
        ["Değişiklik nedeni", result["change_reason"]],
    ]
    story.append(_pdf_table(
        metadata,
        [48 * mm, 215 * mm],
        body_style,
        header_style,
        font_name=regular,
    ))

    story.append(Paragraph("2. Alternatiflerin Genel Sonuçları", section_style))
    alternative_rows = [["Alternatif", "Toplam Puan", "Durum", "Karar Açıklaması"]]
    alternative_statuses: list[str] = []
    for alternative in result["alternatives"]:
        status = _clean(alternative.get("status"))
        if status == STATUS_SUITABLE:
            explanation = "Tüm veriler tamam ve zorunlu kriterler sağlandı."
        elif status == STATUS_UNSUITABLE:
            explanation = "En az bir zorunlu kriter sağlanmadı."
        else:
            explanation = "Eksik veri nedeniyle toplam puan kesinleştirilemedi."
        alternative_rows.append([
            alternative.get("alternative_name"),
            _format_number(alternative.get("total_score"), "/100"),
            status,
            explanation,
        ])
        alternative_statuses.append(status)
    story.append(_pdf_table(
        alternative_rows,
        [55 * mm, 28 * mm, 35 * mm, 145 * mm],
        body_style,
        header_style,
        alternative_statuses,
        regular,
    ))

    story.append(Paragraph("3. Parametre Karşılaştırması ve Fark Analizi", section_style))
    comparison_rows = [[
        "Alternatif",
        "Parametre",
        "Mevcut",
        "Alternatif",
        "Fark",
        "Fark %",
        "Birim",
        "Ağırlık",
        "Yön",
        "Kabul Aralığı",
        "Zorunlu",
        "Puan / Durum",
    ]]
    comparison_statuses: list[str] = []
    for alternative in result["alternatives"]:
        for criterion in alternative.get("criteria", []):
            status = _clean(criterion.get("status"))
            comparison_rows.append([
                alternative.get("alternative_name"),
                criterion.get("parameter_name"),
                _format_number(criterion.get("current_value")),
                _format_number(criterion.get("alternative_value")),
                _format_number(criterion.get("difference")),
                _format_number(criterion.get("difference_percent"), "%"),
                criterion.get("unit") or "-",
                _format_number(criterion.get("normalized_weight"), "%"),
                criterion.get("direction"),
                (
                    f"{_format_number(criterion.get('minimum'))} - "
                    f"{_format_number(criterion.get('maximum'))}"
                ),
                "Evet" if criterion.get("mandatory") else "Hayır",
                (
                    f"{_format_number(criterion.get('criterion_score'))} / "
                    f"{status}"
                ),
            ])
            comparison_statuses.append(status)
    widths = [
        24 * mm, 28 * mm, 17 * mm, 17 * mm, 16 * mm, 16 * mm,
        13 * mm, 16 * mm, 24 * mm, 25 * mm, 15 * mm, 35 * mm,
    ]
    story.append(_pdf_table(
        comparison_rows,
        widths,
        small_style,
        header_style,
        comparison_statuses,
        regular,
    ))

    story.append(Paragraph("4. Zorunlu Kriter Sonuçları", section_style))
    mandatory_rows = [[
        "Alternatif", "Parametre", "Alternatif Değer", "Kabul Aralığı", "Sonuç", "Açıklama"
    ]]
    mandatory_statuses: list[str] = []
    for alternative in result["alternatives"]:
        for criterion in alternative.get("criteria", []):
            if not criterion.get("mandatory"):
                continue
            status = _clean(criterion.get("status"))
            mandatory_rows.append([
                alternative.get("alternative_name"),
                criterion.get("parameter_name"),
                _format_number(criterion.get("alternative_value")),
                (
                    f"{_format_number(criterion.get('minimum'))} - "
                    f"{_format_number(criterion.get('maximum'))} "
                    f"{criterion.get('unit') or ''}"
                ),
                status,
                _criterion_explanation(criterion),
            ])
            mandatory_statuses.append(status)
    if len(mandatory_rows) == 1:
        mandatory_rows.append(["-", "-", "-", "-", "Uyarı", "Zorunlu kriter tanımlanmadı."])
        mandatory_statuses.append("Uyarı")
    story.append(_pdf_table(
        mandatory_rows,
        [38 * mm, 40 * mm, 30 * mm, 42 * mm, 38 * mm, 75 * mm],
        body_style,
        header_style,
        mandatory_statuses,
        regular,
    ))

    story.append(Paragraph("5. Önerilen Alternatif ve Seçim Gerekçesi", section_style))
    best = result.get("best_alternative")
    if best:
        recommendation = Table(
            [[Paragraph(
                _paragraph_text(
                    f"{best['alternative_name']} - {_format_number(best['total_score'])}/100"
                ),
                callout_style,
            )]],
            colWidths=[263 * mm],
        )
        recommendation.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(f"#{SUCCESS}")),
            ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor(f"#{SUCCESS_TEXT}")),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ]))
        story.append(recommendation)
        story.append(Spacer(1, 2 * mm))
        story.append(Paragraph(
            "Bu alternatif, 'Uygun' durumundaki seçenekler arasında en yüksek "
            "kayıtlı toplam puana sahiptir ve zorunlu kriterleri sağlamaktadır.",
            body_style,
        ))
    else:
        story.append(Paragraph(
            "Eksik veriler veya sağlanmayan zorunlu kriterler nedeniyle önerilen "
            "alternatif belirlenemedi.",
            body_style,
        ))

    story.append(Paragraph("6. Eksik Veri ve Uyarılar", section_style))
    warnings = _collect_warnings(result)
    warning_rows = [["Seviye", "Alternatif", "Parametre", "Açıklama"]]
    warning_statuses: list[str] = []
    if warnings:
        for warning in warnings:
            warning_rows.append([
                warning["level"],
                warning["alternative"],
                warning["parameter"],
                warning["message"],
            ])
            warning_statuses.append(warning["level"])
    else:
        warning_rows.append(["Uygun", "-", "-", "Eksik veri veya raporlanacak uyarı bulunmadı."])
        warning_statuses.append("Uygun")
    story.append(_pdf_table(
        warning_rows,
        [30 * mm, 45 * mm, 45 * mm, 143 * mm],
        body_style,
        header_style,
        warning_statuses,
        regular,
    ))

    story.append(PageBreak())
    story.append(Paragraph("7. Hesaplama Yöntemi ve Ağırlıklar", section_style))
    for explanation in result.get("calculation_explanation", []):
        story.append(Paragraph(f"• {_paragraph_text(explanation)}", body_style))
        story.append(Spacer(1, 1.5 * mm))
    weight_rows = [["Parametre", "Normalize Ağırlık"]]
    for parameter_name, weight in result["normalized_weights"].items():
        weight_rows.append([parameter_name, _format_number(weight, "%")])
    story.append(Spacer(1, 2 * mm))
    story.append(_pdf_table(
        weight_rows,
        [180 * mm, 83 * mm],
        body_style,
        header_style,
        font_name=regular,
    ))

    analysis_name = _clean(result["analysis_name"])

    def draw_header_footer(canvas, document) -> None:
        canvas.saveState()
        page_width, page_height = landscape(A4)
        canvas.setStrokeColor(colors.HexColor(f"#{BORDER}"))
        canvas.setLineWidth(0.4)
        canvas.line(12 * mm, page_height - 12 * mm, page_width - 12 * mm, page_height - 12 * mm)
        canvas.setFont(bold, 7)
        canvas.setFillColor(colors.HexColor(f"#{GRAPHITE}"))
        canvas.drawString(12 * mm, page_height - 9.5 * mm, analysis_name[:110])
        canvas.line(12 * mm, 11 * mm, page_width - 12 * mm, 11 * mm)
        canvas.setFont(regular, 7)
        canvas.drawString(
            12 * mm,
            7.5 * mm,
            f"Rapor tarihi: {created_at.strftime('%d.%m.%Y %H:%M')}",
        )
        canvas.drawRightString(
            page_width - 12 * mm,
            7.5 * mm,
            f"Sayfa {document.page}",
        )
        canvas.restoreState()

    document = SimpleDocTemplate(
        str(output_path),
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=18 * mm,
        bottomMargin=17 * mm,
        title=f"{analysis_name} - Etki Analizi Raporu",
        author="Etki Analizi Uygulaması",
        subject="Alternatif karşılaştırma ve etki analizi sonuçları",
    )
    document.build(
        story,
        onFirstPage=draw_header_footer,
        onLaterPages=draw_header_footer,
    )
    return {
        "path": str(output_path.resolve()),
        "format": "pdf",
        "alternative_count": len(result["alternatives"]),
        "warning_count": len(warnings),
        "created_at": created_at,
    }


def _style_title(ws, cell_range: str, title: str) -> None:
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = title
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(name="Segoe UI", size=15, bold=True, color=WHITE)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[cell.row].height = 28


def _style_header(ws, row: int, min_col: int, max_col: int) -> None:
    thin = Side(style="thin", color=BORDER)
    for column in range(min_col, max_col + 1):
        cell = ws.cell(row=row, column=column)
        cell.fill = PatternFill("solid", fgColor=GRAPHITE)
        cell.font = Font(name="Segoe UI", size=9, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.row_dimensions[row].height = 28


def _style_body(ws, min_row: int, max_row: int, max_col: int) -> None:
    if max_row < min_row:
        return
    thin = Side(style="thin", color=BORDER)
    for row in ws.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=1,
        max_col=max_col,
    ):
        for cell in row:
            cell.font = Font(name="Segoe UI", size=9, color="222222")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)


def _add_excel_table(ws, name: str, ref: str) -> None:
    table = ExcelTable(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _fit_columns(ws, min_width: int = 10, max_width: int = 48) -> None:
    for column_index in range(1, ws.max_column + 1):
        letter = get_column_letter(column_index)
        length = 0
        for cell in ws[letter]:
            if cell.value is None:
                continue
            line_length = max(
                (len(part) for part in str(cell.value).splitlines()),
                default=0,
            )
            length = max(length, line_length)
        ws.column_dimensions[letter].width = min(max(length + 2, min_width), max_width)


def _add_status_rules(
    ws,
    cell_range: str,
    anchor_column: str,
    start_row: int = 2,
) -> None:
    rules = (
        (f'${anchor_column}{start_row}="{STATUS_SUITABLE}"', SUCCESS, SUCCESS_TEXT),
        (f'${anchor_column}{start_row}="{STATUS_UNSUITABLE}"', DANGER, DANGER_TEXT),
        (f'${anchor_column}{start_row}="{STATUS_MISSING}"', MISSING, MISSING_TEXT),
        (f'${anchor_column}{start_row}="{STATUS_WARNING}"', WARNING, WARNING_TEXT),
        (f'${anchor_column}{start_row}="{STATUS_MANDATORY_FAIL}"', DANGER, DANGER_TEXT),
        (f'${anchor_column}{start_row}="Olumlu"', SUCCESS, SUCCESS_TEXT),
        (f'${anchor_column}{start_row}="Olumsuz"', DANGER, DANGER_TEXT),
        (f'${anchor_column}{start_row}="Nötr"', WARNING, WARNING_TEXT),
        (f'${anchor_column}{start_row}="Uyarı"', WARNING, WARNING_TEXT),
    )
    for formula, fill, font_color in rules:
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(
                formula=[formula],
                fill=PatternFill("solid", fgColor=fill),
                font=Font(color=font_color),
            ),
        )


def _apply_numeric_format(ws, columns: Sequence[int], start_row: int, end_row: int) -> None:
    for column in columns:
        for row in range(start_row, end_row + 1):
            ws.cell(row=row, column=column).number_format = "0.00"


def _apply_percent_point_format(
    ws,
    columns: Sequence[int],
    start_row: int,
    end_row: int,
) -> None:
    """Yüzde puanı olarak saklanan 25.5 gibi değerleri 25.50% biçiminde gösterir."""
    for column in columns:
        for row in range(start_row, end_row + 1):
            ws.cell(row=row, column=column).number_format = '0.00"%"'


def export_impact_analysis_excel(
    path: str | Path,
    result: Mapping[str, Any],
    report_datetime: datetime | None = None,
) -> dict[str, Any]:
    """Mevcut Etki Analizi sonucundan beş sayfalı Excel raporu üretir."""
    _validate_result(result)
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    created_at = report_datetime or datetime.now()

    workbook = Workbook()
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcOnSave = True
    summary_ws = workbook.active
    summary_ws.title = "Özet"
    parameters_ws = workbook.create_sheet("Parametreler")
    differences_ws = workbook.create_sheet("Fark Analizi")
    criteria_ws = workbook.create_sheet("Kriter Sonuçları")
    calculation_ws = workbook.create_sheet("Hesaplama")
    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False

    _style_title(summary_ws, "A1:F1", "ETKİ ANALİZİ · SONUÇ ÖZETİ")
    summary_info = (
        ("Analiz Adı", result["analysis_name"]),
        ("Mevcut Parça veya Durum", result["current_state"]),
        ("Değişiklik Nedeni", result["change_reason"]),
        ("Rapor Tarihi", created_at),
    )
    for row, (label, value) in enumerate(summary_info, start=3):
        summary_ws.cell(row=row, column=1, value=label)
        summary_ws.cell(row=row, column=2, value=value)
        summary_ws.cell(row=row, column=1).font = Font(
            name="Segoe UI", size=9, bold=True, color=GRAPHITE
        )
    summary_ws["B6"].number_format = "yyyy-mm-dd hh:mm"
    summary_ws.row_dimensions[5].height = 42
    summary_header_row = 9
    summary_ws.cell(summary_header_row - 1, 1, "Alternatiflerin Genel Sonuçları")
    summary_ws.cell(summary_header_row - 1, 1).font = Font(
        name="Segoe UI", size=10, bold=True, color=NAVY
    )
    summary_ws.append([])
    for column, value in enumerate(("Alternatif", "Toplam Puan", "Durum", "Açıklama"), start=1):
        summary_ws.cell(summary_header_row, column, value)
    for alternative in result["alternatives"]:
        status = _clean(alternative.get("status"))
        explanation = (
            "Tüm veriler tamam; zorunlu kriterler sağlandı."
            if status == STATUS_SUITABLE
            else "En az bir zorunlu kriter sağlanmadı."
            if status == STATUS_UNSUITABLE
            else "Eksik veri nedeniyle toplam puan hesaplanamadı."
        )
        summary_ws.append([
            alternative.get("alternative_name"),
            alternative.get("total_score"),
            status,
            explanation,
        ])
    summary_end_row = summary_ws.max_row
    _style_header(summary_ws, summary_header_row, 1, 4)
    _style_body(summary_ws, summary_header_row + 1, summary_end_row, 4)
    _add_excel_table(
        summary_ws,
        "ImpactSummaryTable",
        f"A{summary_header_row}:D{summary_end_row}",
    )
    summary_ws.freeze_panes = f"A{summary_header_row + 1}"
    _apply_numeric_format(summary_ws, (2,), summary_header_row + 1, summary_end_row)
    _add_status_rules(
        summary_ws,
        f"C{summary_header_row + 1}:C{summary_end_row}",
        "C",
        summary_header_row + 1,
    )
    recommended_row = summary_end_row + 3
    summary_ws.cell(recommended_row, 1, "Önerilen Alternatif")
    summary_ws.cell(recommended_row, 1).font = Font(
        name="Segoe UI", size=10, bold=True, color=NAVY
    )
    best = result.get("best_alternative")
    if best:
        summary_ws.cell(recommended_row, 2, best["alternative_name"])
        summary_ws.cell(recommended_row, 3, best["total_score"])
        summary_ws.cell(recommended_row, 3).number_format = "0.00"
        summary_ws.cell(
            recommended_row + 1,
            2,
            "Uygun alternatifler arasındaki en yüksek kayıtlı toplam puan.",
        )
        for cell in summary_ws[recommended_row]:
            cell.fill = PatternFill("solid", fgColor=SUCCESS)
    else:
        summary_ws.cell(recommended_row, 2, "Belirlenemedi")
        summary_ws.cell(
            recommended_row + 1,
            2,
            "Eksik veri veya sağlanmayan zorunlu kriter bulunuyor.",
        )
        for cell in summary_ws[recommended_row]:
            cell.fill = PatternFill("solid", fgColor=WARNING)
    _fit_columns(summary_ws)
    summary_ws.column_dimensions["B"].width = max(summary_ws.column_dimensions["B"].width, 40)
    summary_ws.column_dimensions["D"].width = 52

    parameter_headers = (
        "Alternatif",
        "Parametre Adı",
        "Mevcut Değer",
        "Alternatif Değer",
        "Birim",
        "Ağırlık %",
        "Değer Yönü",
        "Minimum Sınır",
        "Maksimum Sınır",
        "Zorunlu Kriter",
        "Parametre Puanı",
        "Durum",
    )
    parameters_ws.append(parameter_headers)
    for alternative in result["alternatives"]:
        for criterion in alternative.get("criteria", []):
            parameters_ws.append([
                alternative.get("alternative_name"),
                criterion.get("parameter_name"),
                criterion.get("current_value"),
                criterion.get("alternative_value"),
                criterion.get("unit"),
                criterion.get("normalized_weight"),
                criterion.get("direction"),
                criterion.get("minimum"),
                criterion.get("maximum"),
                "Evet" if criterion.get("mandatory") else "Hayır",
                criterion.get("criterion_score"),
                criterion.get("status"),
            ])
    parameter_end = parameters_ws.max_row
    _style_header(parameters_ws, 1, 1, len(parameter_headers))
    _style_body(parameters_ws, 2, parameter_end, len(parameter_headers))
    _add_excel_table(parameters_ws, "ImpactParametersTable", f"A1:L{parameter_end}")
    parameters_ws.freeze_panes = "A2"
    _apply_numeric_format(parameters_ws, (3, 4, 6, 8, 9, 11), 2, parameter_end)
    _apply_percent_point_format(parameters_ws, (6,), 2, parameter_end)
    _add_status_rules(parameters_ws, f"L2:L{parameter_end}", "L")
    _fit_columns(parameters_ws)

    difference_headers = (
        "Alternatif",
        "Parametre",
        "Mevcut Değer",
        "Alternatif Değer",
        "Mutlak Değişim",
        "Yüzde Değişim",
        "Birim",
        "Değer Yönü",
        "Değerlendirme",
    )
    differences_ws.append(difference_headers)
    for alternative in result["alternatives"]:
        for criterion in alternative.get("criteria", []):
            differences_ws.append([
                alternative.get("alternative_name"),
                criterion.get("parameter_name"),
                criterion.get("current_value"),
                criterion.get("alternative_value"),
                criterion.get("difference"),
                criterion.get("difference_percent"),
                criterion.get("unit"),
                criterion.get("direction"),
                _difference_evaluation(criterion),
            ])
    difference_end = differences_ws.max_row
    _style_header(differences_ws, 1, 1, len(difference_headers))
    _style_body(differences_ws, 2, difference_end, len(difference_headers))
    _add_excel_table(differences_ws, "ImpactDifferencesTable", f"A1:I{difference_end}")
    differences_ws.freeze_panes = "A2"
    _apply_numeric_format(differences_ws, (3, 4, 5, 6), 2, difference_end)
    _apply_percent_point_format(differences_ws, (6,), 2, difference_end)
    _add_status_rules(differences_ws, f"I2:I{difference_end}", "I")
    _fit_columns(differences_ws)

    criteria_headers = (
        "Alternatif",
        "Parametre",
        "Zorunlu Kriter",
        "Alternatif Değer",
        "Kabul Aralığı",
        "Durum",
        "Açıklama",
    )
    criteria_ws.append(criteria_headers)
    mandatory_count = 0
    for alternative in result["alternatives"]:
        for criterion in alternative.get("criteria", []):
            if not criterion.get("mandatory"):
                continue
            mandatory_count += 1
            criteria_ws.append([
                alternative.get("alternative_name"),
                criterion.get("parameter_name"),
                "Evet",
                criterion.get("alternative_value"),
                (
                    f"{_format_number(criterion.get('minimum'))} - "
                    f"{_format_number(criterion.get('maximum'))} "
                    f"{criterion.get('unit') or ''}"
                ),
                criterion.get("status"),
                _criterion_explanation(criterion),
            ])
    if mandatory_count == 0:
        criteria_ws.append([
            "-", "Zorunlu kriter tanımlanmadı", "Hayır", None, "-", "Uyarı",
            "Analizde zorunlu olarak işaretlenmiş bir kriter bulunmuyor.",
        ])
    criteria_end = criteria_ws.max_row
    _style_header(criteria_ws, 1, 1, len(criteria_headers))
    _style_body(criteria_ws, 2, criteria_end, len(criteria_headers))
    _add_excel_table(criteria_ws, "ImpactCriteriaTable", f"A1:G{criteria_end}")
    criteria_ws.freeze_panes = "A2"
    _apply_numeric_format(criteria_ws, (4,), 2, criteria_end)
    _add_status_rules(criteria_ws, f"F2:F{criteria_end}", "F")
    _fit_columns(criteria_ws)
    criteria_ws.column_dimensions["G"].width = 52

    _style_title(calculation_ws, "A1:F1", "HESAPLAMA YÖNTEMİ VE KAYITLI SONUÇLAR")
    calculation_ws["A3"] = "Hesaplama Yöntemi"
    calculation_ws["A3"].font = Font(name="Segoe UI", size=10, bold=True, color=NAVY)
    method_start = 4
    for offset, explanation in enumerate(result.get("calculation_explanation", [])):
        calculation_ws.cell(method_start + offset, 1, f"• {explanation}")
        calculation_ws.merge_cells(
            start_row=method_start + offset,
            start_column=1,
            end_row=method_start + offset,
            end_column=6,
        )
        calculation_ws.cell(method_start + offset, 1).alignment = Alignment(wrap_text=True)
    weight_header = method_start + len(result.get("calculation_explanation", [])) + 2
    calculation_ws.cell(weight_header - 1, 1, "Normalize Edilmiş Ağırlıklar")
    calculation_ws.cell(weight_header - 1, 1).font = Font(
        name="Segoe UI", size=10, bold=True, color=NAVY
    )
    calculation_ws.cell(weight_header, 1, "Parametre")
    calculation_ws.cell(weight_header, 2, "Normalize Ağırlık %")
    for parameter_name, weight in result["normalized_weights"].items():
        calculation_ws.append([parameter_name, weight])
    weight_end = calculation_ws.max_row
    _style_header(calculation_ws, weight_header, 1, 2)
    _style_body(calculation_ws, weight_header + 1, weight_end, 2)
    _add_excel_table(
        calculation_ws,
        "ImpactWeightsTable",
        f"A{weight_header}:B{weight_end}",
    )
    _apply_numeric_format(calculation_ws, (2,), weight_header + 1, weight_end)
    _apply_percent_point_format(calculation_ws, (2,), weight_header + 1, weight_end)

    score_header = weight_end + 3
    score_headers = (
        "Alternatif",
        "Parametre",
        "Parametre Puanı",
        "Ağırlık %",
        "Ağırlıklı Katkı",
        "Kayıtlı Toplam Puan",
    )
    for column, value in enumerate(score_headers, start=1):
        calculation_ws.cell(score_header, column, value)
    for alternative in result["alternatives"]:
        for criterion in alternative.get("criteria", []):
            score = criterion.get("criterion_score")
            weight = criterion.get("normalized_weight")
            contribution = (
                None if score is None or weight is None
                else round(float(score) * float(weight) / 100.0, 4)
            )
            calculation_ws.append([
                alternative.get("alternative_name"),
                criterion.get("parameter_name"),
                score,
                weight,
                contribution,
                alternative.get("total_score"),
            ])
    score_end = calculation_ws.max_row
    _style_header(calculation_ws, score_header, 1, len(score_headers))
    _style_body(calculation_ws, score_header + 1, score_end, len(score_headers))
    _add_excel_table(
        calculation_ws,
        "ImpactScoreBreakdownTable",
        f"A{score_header}:F{score_end}",
    )
    _apply_numeric_format(
        calculation_ws,
        (3, 4, 5, 6),
        score_header + 1,
        score_end,
    )
    _apply_percent_point_format(calculation_ws, (4,), score_header + 1, score_end)
    calculation_ws.freeze_panes = f"A{score_header + 1}"
    _fit_columns(calculation_ws)
    calculation_ws.column_dimensions["A"].width = max(
        calculation_ws.column_dimensions["A"].width, 24
    )

    workbook.save(output_path)
    return {
        "path": str(output_path.resolve()),
        "format": "excel",
        "alternative_count": len(result["alternatives"]),
        "warning_count": len(_collect_warnings(result)),
        "sheet_count": len(workbook.sheetnames),
        "created_at": created_at,
    }

